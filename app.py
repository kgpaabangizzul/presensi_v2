from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, g
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
import os, math, json, io, re, random, string, smtplib, ssl, urllib.request as _urllib_req
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps
from collections import OrderedDict
import psycopg2
import psycopg2.extras
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from lupa_password import lupa_pw_bp, init_reset_table
from pengumuman import pengumuman_bp, init_pengumuman_table

app = Flask(__name__)
app.register_blueprint(lupa_pw_bp)
app.register_blueprint(pengumuman_bp)
app.secret_key = 'absensi-secret-key-2024'

@app.before_request
def load_global_settings():
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM settings WHERE id=1")
        row = cur.fetchone()
        g.settings = dict(row) if row else {}
        cur.close(); conn.close()
    except Exception:
        g.settings = {}
app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads', 'photos')
app.config['DOSIR_FOLDER']  = os.path.join('static', 'uploads', 'dosir')
app.config['SURAT_FOLDER']  = os.path.join('static', 'uploads', 'surat')
app.config['TTD_FOLDER']    = os.path.join('static', 'uploads', 'ttd')
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_LAMPIRAN   = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
ALLOWED_DOSIR      = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}

# ── PostgreSQL config ─────────────────────────────────────────────────────────
DB_HOST = os.environ.get('DB_HOST', 'localhost')
DB_PORT = os.environ.get('DB_PORT', '5432')
DB_NAME = os.environ.get('DB_NAME', 'presensi')
DB_USER = os.environ.get('DB_USER', 'presensi')
DB_PASS = os.environ.get('DB_PASS', 'presensi123')

def get_db():
    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT, dbname=DB_NAME,
        user=DB_USER, password=DB_PASS
    )
    conn.autocommit = False
    return conn

def fetchone(cur):
    row = cur.fetchone()
    return dict(row) if row else None

def fetchall(cur):
    return [dict(r) for r in cur.fetchall()]

# ══════════════════════════════════════════════════════════════════════════════
# ██  AUDIT LOG SYSTEM  ████████████████████████████████████████████████████████
# ══════════════════════════════════════════════════════════════════════════════

AKSI_LABELS = {
    'LOGIN'         : ('🔐', 'Login'),
    'LOGOUT'        : ('🚪', 'Logout'),
    'LOGIN_GAGAL'   : ('❌', 'Login Gagal'),
    'REGISTER'      : ('📝', 'Registrasi'),
    'CREATE'        : ('➕', 'Tambah Data'),
    'UPDATE'        : ('✏️',  'Ubah Data'),
    'DELETE'        : ('🗑️',  'Hapus Data'),
    'APPROVE'       : ('✅', 'Setujui'),
    'REJECT'        : ('🚫', 'Tolak'),
    'ABSEN_MASUK'   : ('🟢', 'Absen Masuk'),
    'ABSEN_KELUAR'  : ('🔴', 'Absen Keluar'),
    'EXPORT'        : ('📥', 'Export Data'),
    'UPLOAD'        : ('📤', 'Upload File'),
    'SETTING'       : ('⚙️',  'Ubah Pengaturan'),
    'VIEW'          : ('👁️',  'Lihat Data'),
    'PASSWORD'      : ('🔑', 'Ubah Password'),
    'VALIDASI'      : ('🪪', 'Validasi Akun'),
    'LUPA_ABSEN'    : ('⏱️',  'Lupa Absen'),
    'NOTA_DINAS'    : ('📄', 'Nota Dinas'),
    'SURAT'         : ('📃', 'Surat Perintah'),
    'IZIN'          : ('📋', 'Izin'),
    'OTP_KIRIM'     : ('📨', 'Kirim OTP'),
    'OTP_VERIF'     : ('🔏', 'Verifikasi OTP'),
    'RESET_PW'      : ('🔑', 'Reset Password'),
}

MODUL_LABELS = {
    'auth'          : 'Autentikasi',
    'absensi'       : 'Absensi',
    'izin'          : 'Perizinan',
    'pegawai'       : 'Data Pegawai',
    'departemen'    : 'Departemen',
    'shift'         : 'Shift',
    'settings'      : 'Pengaturan Sistem',
    'laporan'       : 'Laporan',
    'arsip'         : 'Arsip Surat',
    'surat'         : 'Surat Perintah',
    'dosir'         : 'E-Dosir',
    'profil'        : 'Profil',
    'role'          : 'Manajemen Role',
    'permission'    : 'Hak Akses',
    'notifikasi'    : 'Notifikasi',
    'approval'      : 'Approval',
    'sistem'        : 'Sistem',
}


def _init_audit_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id          BIGSERIAL PRIMARY KEY,
            waktu       TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            user_id     INTEGER,
            user_nama   TEXT,
            user_role   TEXT,
            aksi        TEXT NOT NULL,
            modul       TEXT NOT NULL,
            deskripsi   TEXT,
            data_lama   JSONB,
            data_baru   JSONB,
            ref_id      INTEGER,
            ref_table   TEXT,
            ip_address  TEXT,
            user_agent  TEXT,
            status      TEXT DEFAULT 'success',
            pesan_error TEXT
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_waktu ON audit_log(waktu DESC)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_user  ON audit_log(user_id)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_modul ON audit_log(modul)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_audit_aksi  ON audit_log(aksi)")


def log_audit(conn, aksi, modul, deskripsi=None,
              data_lama=None, data_baru=None,
              ref_id=None, ref_table=None,
              status='success', pesan_error=None,
              user_id=None, user_nama=None, user_role=None):
    """Catat satu baris audit log. Aman dipanggil dari mana saja."""
    try:
        if user_id   is None: user_id   = session.get('user_id')
        if user_nama is None: user_nama = session.get('nama', 'System')
        if user_role is None: user_role = session.get('role', '-')
        try:
            ip = request.headers.get('X-Forwarded-For', request.remote_addr) or '-'
            ua = (request.user_agent.string or '-')[:300]
        except RuntimeError:
            ip, ua = 'system', 'system'

        def _sanitize(d):
            if not isinstance(d, dict): return d
            skip = {'password', 'password_hash', 'token', 'secret', 'otp_code'}
            return {k: '***' if k in skip else v for k, v in d.items()}

        dl = json.dumps(_sanitize(data_lama), default=str) if data_lama else None
        db_ = json.dumps(_sanitize(data_baru), default=str) if data_baru else None

        cur = conn.cursor()
        cur.execute("""
            INSERT INTO audit_log
              (user_id,user_nama,user_role,aksi,modul,deskripsi,
               data_lama,data_baru,ref_id,ref_table,
               ip_address,user_agent,status,pesan_error)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (user_id, user_nama, user_role,
              aksi, modul, deskripsi,
              dl, db_, ref_id, ref_table,
              ip, ua, status, pesan_error))
        conn.commit()
        cur.close()
    except Exception as e:
        try: conn.rollback()
        except Exception: pass


def log_error(conn, aksi, modul, pesan_error, deskripsi=None, ref_id=None):
    log_audit(conn, aksi, modul, deskripsi=deskripsi,
              ref_id=ref_id, status='error', pesan_error=str(pesan_error))


# ══════════════════════════════════════════════════════════════════════════════
# ██  LUPA PASSWORD — OTP via Email & WhatsApp  ████████████████████████████████
# ══════════════════════════════════════════════════════════════════════════════

# ── Konfigurasi Notifikasi ────────────────────────────────────────────────────
def _get_notif_config(conn=None):
    cfg = {
        'smtp_host'      : os.environ.get('SMTP_HOST', ''),
        'smtp_port'      : int(os.environ.get('SMTP_PORT', 587)),
        'smtp_user'      : os.environ.get('SMTP_USER', ''),
        'smtp_pass'      : os.environ.get('SMTP_PASS', ''),
        'smtp_from_name' : os.environ.get('SMTP_FROM_NAME', 'Presensi Digital'),
        'smtp_tls'       : os.environ.get('SMTP_TLS', 'true').lower() == 'true',
        'fonnte_token'   : os.environ.get('FONNTE_TOKEN', ''),
        'fonnte_url'     : 'https://api.fonnte.com/send',
        'otp_expire'     : int(os.environ.get('OTP_EXPIRE_MENIT', 10)),
        'otp_length'     : int(os.environ.get('OTP_LENGTH', 6)),
        'nama_perusahaan': 'Presensi Digital',
    }
    try:
        _conn = conn or get_db()
        cur2 = _conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur2.execute("SELECT * FROM settings WHERE id=1")
        row = cur2.fetchone()
        if row:
            cfg['nama_perusahaan'] = row.get('nama_perusahaan', cfg['nama_perusahaan'])
            for k in ['smtp_host','smtp_port','smtp_user','smtp_pass',
                      'smtp_from_name','fonnte_token']:
                if row.get(k): cfg[k] = row[k]
        cur2.close()
        if not conn: _conn.close()
    except Exception: pass
    return cfg


def _init_reset_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS password_reset_otp (
            id         SERIAL PRIMARY KEY,
            user_id    INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            otp_code   TEXT NOT NULL,
            metode     TEXT NOT NULL DEFAULT 'email',
            tujuan     TEXT NOT NULL,
            kadaluarsa TIMESTAMP NOT NULL,
            digunakan  BOOLEAN DEFAULT FALSE,
            ip_address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_otp_user ON password_reset_otp(user_id,digunakan)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_otp_kode ON password_reset_otp(otp_code,digunakan)")
    for kolom, tipe in [
        ('smtp_host',      "TEXT DEFAULT ''"),
        ('smtp_port',      "INTEGER DEFAULT 587"),
        ('smtp_user',      "TEXT DEFAULT ''"),
        ('smtp_pass',      "TEXT DEFAULT ''"),
        ('smtp_from_name', "TEXT DEFAULT 'Presensi Digital'"),
        ('smtp_tls',       "BOOLEAN DEFAULT TRUE"),
        ('fonnte_token',   "TEXT DEFAULT ''"),
    ]:
        try: cur.execute(f"ALTER TABLE settings ADD COLUMN IF NOT EXISTS {kolom} {tipe}")
        except Exception: pass


def _generate_otp(length=6):
    return ''.join(random.choices(string.digits, k=length))


def _simpan_otp(conn, user_id, otp_code, metode, tujuan, expire_menit=10):
    cur = conn.cursor()
    cur.execute("UPDATE password_reset_otp SET digunakan=TRUE WHERE user_id=%s AND digunakan=FALSE", (user_id,))
    kadaluarsa = datetime.now() + timedelta(minutes=expire_menit)
    try: ip = request.remote_addr
    except RuntimeError: ip = 'system'
    cur.execute("""INSERT INTO password_reset_otp (user_id,otp_code,metode,tujuan,kadaluarsa,ip_address)
        VALUES (%s,%s,%s,%s,%s,%s)""", (user_id, otp_code, metode, tujuan, kadaluarsa, ip))
    conn.commit(); cur.close()


def _verifikasi_otp(conn, user_id, otp_input):
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT * FROM password_reset_otp
        WHERE user_id=%s AND otp_code=%s AND digunakan=FALSE AND kadaluarsa>NOW()
        ORDER BY created_at DESC LIMIT 1""", (user_id, otp_input.strip()))
    row = cur.fetchone(); cur.close()
    return dict(row) if row else None


def _tandai_otp_digunakan(conn, otp_id):
    cur = conn.cursor()
    cur.execute("UPDATE password_reset_otp SET digunakan=TRUE WHERE id=%s", (otp_id,))
    conn.commit(); cur.close()


def _kirim_email_otp(cfg, tujuan_email, nama_user, otp_code):
    if not cfg['smtp_host'] or not cfg['smtp_user']:
        return False, "Konfigurasi SMTP belum diisi. Hubungi administrator."
    subject = f"[{cfg['nama_perusahaan']}] Kode OTP Reset Password"
    expire  = cfg['otp_expire']
    html_body = f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="font-family:Arial,sans-serif;background:#f8fafc;margin:0;padding:20px">
  <div style="max-width:480px;margin:0 auto;background:#fff;border-radius:12px;
              box-shadow:0 2px 12px rgba(0,0,0,.08);overflow:hidden">
    <div style="background:linear-gradient(135deg,#2563eb,#1d4ed8);padding:28px 32px">
      <h2 style="color:#fff;margin:0;font-size:20px">🔐 Reset Password</h2>
      <p style="color:rgba(255,255,255,.8);margin:6px 0 0;font-size:13px">{cfg['nama_perusahaan']}</p>
    </div>
    <div style="padding:32px">
      <p style="color:#1e293b;margin:0 0 16px">Halo <strong>{nama_user}</strong>,</p>
      <p style="color:#475569;font-size:14px;margin:0 0 24px">
        Kami menerima permintaan reset password. Gunakan kode OTP berikut:
      </p>
      <div style="background:#f1f5f9;border:2px dashed #cbd5e1;border-radius:10px;
                  padding:20px;text-align:center;margin:0 0 24px">
        <div style="font-size:38px;font-weight:700;letter-spacing:10px;
                    color:#2563eb;font-family:monospace">{otp_code}</div>
        <div style="color:#94a3b8;font-size:12px;margin-top:8px">
          ⏱ Berlaku selama <strong>{expire} menit</strong>
        </div>
      </div>
      <div style="background:#fef9c3;border-left:4px solid #eab308;
                  padding:12px 16px;border-radius:0 8px 8px 0;margin:0 0 24px">
        <p style="margin:0;font-size:13px;color:#713f12">
          ⚠️ <strong>Jangan bagikan kode ini</strong> kepada siapapun.
          Jika tidak merasa meminta reset, abaikan email ini.
        </p>
      </div>
    </div>
  </div>
</body></html>"""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f"{cfg['smtp_from_name']} <{cfg['smtp_user']}>"
        msg['To']      = tujuan_email
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        port = int(cfg['smtp_port'])
        if cfg['smtp_tls']:
            ctx = ssl.create_default_context()
            with smtplib.SMTP(cfg['smtp_host'], port, timeout=15) as srv:
                srv.ehlo(); srv.starttls(context=ctx)
                srv.login(cfg['smtp_user'], cfg['smtp_pass'])
                srv.sendmail(cfg['smtp_user'], tujuan_email, msg.as_string())
        else:
            with smtplib.SMTP_SSL(cfg['smtp_host'], port, timeout=15) as srv:
                srv.login(cfg['smtp_user'], cfg['smtp_pass'])
                srv.sendmail(cfg['smtp_user'], tujuan_email, msg.as_string())
        return True, ''
    except smtplib.SMTPAuthenticationError:
        return False, "Autentikasi SMTP gagal. Periksa email & App Password."
    except Exception as e:
        return False, f"Gagal kirim email: {str(e)}"


def _kirim_wa_otp(cfg, no_hp, nama_user, otp_code):
    if not cfg['fonnte_token']:
        return False, "Token Fonnte WhatsApp belum dikonfigurasi. Hubungi administrator."
    nomor = re.sub(r'\D', '', no_hp)
    if nomor.startswith('0'): nomor = '62' + nomor[1:]
    elif not nomor.startswith('62'): nomor = '62' + nomor
    expire = cfg['otp_expire']
    pesan = (f"🔐 *Reset Password — {cfg['nama_perusahaan']}*\n\n"
             f"Halo *{nama_user}*,\n\n"
             f"Kode OTP reset password Anda:\n\n"
             f"*{otp_code}*\n\n"
             f"⏱ Berlaku {expire} menit.\n\n"
             f"⚠️ Jangan bagikan kode ini kepada siapapun.")
    try:
        payload = json.dumps({'target':nomor,'message':pesan,'countryCode':'62'}).encode('utf-8')
        req = _urllib_req.Request(cfg['fonnte_url'], data=payload,
            headers={'Authorization':cfg['fonnte_token'],'Content-Type':'application/json'},
            method='POST')
        with _urllib_req.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode())
            if result.get('status') in (True, 'true'): return True, ''
            return False, f"Fonnte: {result.get('reason') or result.get('message') or str(result)}"
    except Exception as e:
        return False, f"Gagal kirim WhatsApp: {str(e)}"


def _mask_tujuan(value, metode):
    if not value: return '***'
    if metode == 'email':
        parts = value.split('@')
        if len(parts) == 2:
            name, domain = parts
            return name[:2] + '*'*max(1,len(name)-2) + '@' + domain
        return value[:3] + '***'
    digits = re.sub(r'\D', '', value)
    return digits[:4]+'****'+digits[-3:] if len(digits)>=8 else '****'



def init_db():
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS departemen (
            id SERIAL PRIMARY KEY,
            nama TEXT UNIQUE NOT NULL,
            kode TEXT UNIQUE NOT NULL,
            deskripsi TEXT,
            warna TEXT DEFAULT '#3b82f6',
            aktif INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS shift (
            id SERIAL PRIMARY KEY,
            nama TEXT NOT NULL UNIQUE,
            jam_masuk TEXT NOT NULL,
            jam_keluar TEXT NOT NULL,
            toleransi_menit INTEGER DEFAULT 15,
            deskripsi TEXT,
            warna TEXT DEFAULT '#10b981',
            aktif INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Tambahkan UNIQUE constraint pada kolom nama jika belum ada (untuk DB yang sudah ada)
    try:
        cur.execute("CREATE UNIQUE INDEX IF NOT EXISTS shift_nama_unique ON shift(nama)")
        conn.commit()
    except Exception:
        conn.rollback()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS departemen_shift (
            id SERIAL PRIMARY KEY,
            departemen_id INTEGER NOT NULL REFERENCES departemen(id),
            shift_id INTEGER NOT NULL REFERENCES shift(id),
            UNIQUE(departemen_id, shift_id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            nik TEXT UNIQUE NOT NULL,
            nama TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            jabatan TEXT,
            departemen TEXT,
            departemen_id INTEGER REFERENCES departemen(id),
            shift_id INTEGER REFERENCES shift(id),
            no_hp TEXT,
            alamat TEXT,
            tanggal_lahir TEXT,
            jenis_kelamin TEXT,
            foto TEXT,
            role TEXT DEFAULT 'user',
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS absensi (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            tanggal DATE NOT NULL,
            jam_masuk TEXT,
            jam_keluar TEXT,
            foto_masuk TEXT,
            foto_keluar TEXT,
            lat_masuk REAL,
            lng_masuk REAL,
            lat_keluar REAL,
            lng_keluar REAL,
            jarak_masuk REAL,
            jarak_keluar REAL,
            shift_id INTEGER,
            status TEXT DEFAULT 'hadir',
            keterangan TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS izin (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id),
            tanggal_mulai DATE NOT NULL,
            tanggal_selesai DATE NOT NULL,
            jenis TEXT NOT NULL,
            alasan TEXT,
            lampiran TEXT,
            status TEXT DEFAULT 'pending',
            catatan_admin TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            id INTEGER PRIMARY KEY,
            nama_perusahaan TEXT DEFAULT 'PT. Absensi Digital',
            jam_masuk TEXT DEFAULT '08:00',
            jam_keluar TEXT DEFAULT '17:00',
            office_lat REAL DEFAULT -6.2088,
            office_lng REAL DEFAULT 106.8456,
            max_distance INTEGER DEFAULT 100
        )
    """)

    cur.execute("INSERT INTO settings (id, nama_perusahaan) VALUES (1, 'PT. Absensi Digital') ON CONFLICT DO NOTHING")

    # ── MASTER ROLE ───────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS master_role (
            id SERIAL PRIMARY KEY,
            kode TEXT UNIQUE NOT NULL,
            nama TEXT NOT NULL,
            deskripsi TEXT,
            aktif INTEGER DEFAULT 1,
            urutan INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Seed default roles
    default_roles = [
        ('admin',    'Administrator',  'Akses penuh ke semua fitur', 0),
        ('user',     'Pegawai',        'Akses standar pegawai',      1),
        ('manajer',  'Manajer',        'Akses laporan dan approval', 2),
        ('dokter',   'Dokter',         'Tenaga medis dokter',        3),
        ('perawat',  'Perawat',        'Tenaga medis perawat',       4),
        ('apoteker', 'Apoteker',       'Tenaga farmasi',             5),
        ('bidan',    'Bidan',          'Tenaga kebidanan',           6),
        ('teknisi',  'Teknisi',        'Teknisi dan IT',             7),
        ('security', 'Security',       'Petugas keamanan',           8),
        ('staf',     'Staf Admin',     'Staf administrasi',          9),
    ]
    for kode, nama, desk, urut in default_roles:
        cur.execute("""INSERT INTO master_role (kode,nama,deskripsi,urutan)
            VALUES (%s,%s,%s,%s) ON CONFLICT (kode) DO NOTHING""",
            (kode, nama, desk, urut))

    # ── KOLOM NIP DI USERS ────────────────────────────────────────────────────
    try:
        cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS nip TEXT")
        conn.commit()
    except Exception:
        conn.rollback()

    # ── ARSIP SURAT ───────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS arsip (
            id SERIAL PRIMARY KEY,
            judul TEXT NOT NULL,
            jenis TEXT NOT NULL,
            nomor TEXT,
            tanggal DATE DEFAULT CURRENT_DATE,
            keterangan TEXT,
            filename TEXT,
            original_name TEXT,
            dibuat_oleh INTEGER REFERENCES users(id),
            aktif INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    try:
        cur.execute("CREATE INDEX IF NOT EXISTS idx_arsip_jenis ON arsip(jenis)")
        cur.execute("CREATE INDEX IF NOT EXISTS idx_arsip_tanggal ON arsip(tanggal DESC)")
        conn.commit()
    except Exception:
        conn.rollback()
    # Kolom tambahan arsip (kategori & sumber)
    for _col, _def in [
        ('kategori', 'TEXT'),
        ('sumber',   "TEXT DEFAULT 'admin'"),
    ]:
        try:
            cur.execute(f"ALTER TABLE arsip ADD COLUMN IF NOT EXISTS {_col} {_def}")
            conn.commit()
        except Exception:
            conn.rollback()
    # Tabel distribusi surat ke user
    cur.execute("""
        CREATE TABLE IF NOT EXISTS arsip_bagikan (
            id SERIAL PRIMARY KEY,
            arsip_id INTEGER NOT NULL REFERENCES arsip(id) ON DELETE CASCADE,
            user_id  INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            wa_sent  INTEGER DEFAULT 0,
            wa_error TEXT,
            sent_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(arsip_id, user_id)
        )
    """)
    try:
        conn.commit()
    except Exception:
        conn.rollback()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS notifikasi (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            judul TEXT NOT NULL,
            pesan TEXT,
            tipe TEXT DEFAULT 'info',
            ref_id INTEGER,
            ref_type TEXT,
            dibaca INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Tambah kolom logo jika belum ada
    try:
        cur.execute("ALTER TABLE settings ADD COLUMN IF NOT EXISTS logo TEXT")
    except Exception:
        pass


    # ── HAK AKSES ROLE ────────────────────────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS role_permission (
            id SERIAL PRIMARY KEY,
            role_kode TEXT NOT NULL,
            modul_kode TEXT NOT NULL,
            modul_nama TEXT NOT NULL,
            grup TEXT NOT NULL DEFAULT 'Lainnya',
            aktif INTEGER DEFAULT 1,
            UNIQUE(role_kode, modul_kode)
        )
    """)

    SEMUA_MODUL = [
        ('dashboard',         'Dashboard',                 'Dashboard & Profil'),
        ('profil',            'Profil & Ubah Password',    'Dashboard & Profil'),
        ('riwayat',           'Riwayat Absensi',           'Dashboard & Profil'),
        ('absen',             'Absen Masuk / Keluar',      'Absensi'),
        ('lupa_absen',        'Lapor Lupa Absen',          'Absensi'),
        ('izin',              'Pengajuan Izin',            'Absensi'),
        ('arsip',             'Lihat Arsip Surat',         'Arsip'),
        ('dosir',             'E-Dosir (Upload Dokumen)',  'E-Dosir'),
        ('admin_dashboard',   'Dashboard Admin',           'Admin — Umum'),
        ('admin_pegawai',     'Kelola Pegawai',            'Admin — Umum'),
        ('admin_validasi',    'Validasi Registrasi',       'Admin — Umum'),
        ('admin_settings',    'Pengaturan Sistem',         'Admin — Umum'),
        ('admin_absensi',     'Monitor Absensi',           'Admin — Absensi'),
        ('admin_izin',        'Kelola Izin',               'Admin — Absensi'),
        ('admin_laporan',     'Laporan & Export',          'Admin — Absensi'),
        ('admin_grafik',      'Grafik Statistik',          'Admin — Absensi'),
        ('admin_departemen',  'Kelola Departemen',         'Admin — Master Data'),
        ('admin_shift',       'Kelola Shift',              'Admin — Master Data'),
        ('admin_master_role', 'Kelola Role',               'Admin — Master Data'),
        ('admin_role_perm',   'Kelola Hak Akses',          'Admin — Master Data'),
        ('admin_arsip',       'Kelola Arsip Surat',        'Admin — Arsip'),
        ('admin_dosir',       'Kelola E-Dosir',            'Admin — E-Dosir'),
        ('admin_dosir_jenis', 'Jenis Dokumen Dosir',       'Admin — E-Dosir'),
    ]

    DEFAULT_PERMS = {
        'admin':    [m[0] for m in SEMUA_MODUL],
        'user':     ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir'],
        'manajer':  ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir',
                     'admin_dashboard','admin_absensi','admin_laporan','admin_grafik',
                     'admin_izin','admin_arsip'],
        'dokter':   ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir'],
        'perawat':  ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir'],
        'apoteker': ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir'],
        'bidan':    ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir'],
        'teknisi':  ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir'],
        'security': ['dashboard','profil','riwayat','absen','lupa_absen','izin'],
        'staf':     ['dashboard','profil','riwayat','absen','lupa_absen','izin',
                     'arsip','dosir'],
    }

    for kode, nama, grup in SEMUA_MODUL:
        for role_kode, allowed in DEFAULT_PERMS.items():
            aktif = 1 if kode in allowed else 0
            cur.execute("""INSERT INTO role_permission (role_kode,modul_kode,modul_nama,grup,aktif)
                VALUES (%s,%s,%s,%s,%s) ON CONFLICT (role_kode,modul_kode) DO NOTHING""",
                (role_kode, kode, nama, grup, aktif))

    # ── WEBAUTHN CREDENTIALS ─────────────────────────────────────
    cur.execute("""
        CREATE TABLE IF NOT EXISTS webauthn_credentials (
            id SERIAL PRIMARY KEY,
            user_id INTEGER NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            credential_id TEXT UNIQUE NOT NULL,
            public_key TEXT NOT NULL,
            sign_count BIGINT DEFAULT 0,
            device_name TEXT DEFAULT 'Perangkat',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            last_used TIMESTAMP
        )
    """)
    try:
        cur.execute("CREATE INDEX IF NOT EXISTS idx_wac_user ON webauthn_credentials(user_id)")
        conn.commit()
    except Exception:
        conn.rollback()

    # ── AUDIT LOG TABLE ───────────────────────────────────────────
    _init_audit_table(cur)
    # ── OTP RESET PASSWORD TABLE ──────────────────────────────────
    _init_reset_table(cur)
    # ── PENGUMUMAN TABLE ──────────────────────────────────────────
    init_pengumuman_table(cur)
    # ─────────────────────────────────────────────────────────────

    conn.commit()
    cur.close()
    conn.close()
    print("Database initialized.")

def allowed_file(fn):
    return '.' in fn and fn.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_lampiran(fn):
    """FIX: cek ekstensi untuk lampiran izin (termasuk PDF)"""
    return '.' in fn and fn.rsplit('.', 1)[1].lower() in ALLOWED_LAMPIRAN

def haversine(lat1, lon1, lat2, lon2):
    R = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2 - lat1)
    dl = math.radians(lon2 - lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return 2 * R * math.atan2(math.sqrt(a), math.sqrt(1-a))

def get_user_shift(uid, conn):
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE id=%s", (uid,))
    user = cur.fetchone()
    if user and user['shift_id']:
        cur.execute("SELECT * FROM shift WHERE id=%s AND aktif=1", (user['shift_id'],))
        s = cur.fetchone()
        if s: cur.close(); return dict(s)
    if user and user['departemen_id']:
        cur.execute("""SELECT s.* FROM shift s JOIN departemen_shift ds ON s.id=ds.shift_id
            WHERE ds.departemen_id=%s AND s.aktif=1 LIMIT 1""", (user['departemen_id'],))
        s = cur.fetchone()
        if s: cur.close(); return dict(s)
    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    cur.close()
    return {'jam_masuk': settings['jam_masuk'] if settings else '08:00',
            'jam_keluar': settings['jam_keluar'] if settings else '17:00',
            'toleransi_menit': 15, 'nama': 'Default', 'id': None}

def get_active_shift(uid, conn):
    """Deteksi shift yang sedang berjalan berdasarkan jam sekarang."""
    now = datetime.now().strftime('%H:%M')
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE id=%s", (uid,))
    user = cur.fetchone()
    if user and user['departemen_id']:
        cur.execute("""SELECT s.* FROM shift s
            JOIN departemen_shift ds ON s.id=ds.shift_id
            WHERE ds.departemen_id=%s AND s.aktif=1 ORDER BY s.jam_masuk""", (user['departemen_id'],))
        shifts = cur.fetchall()
    else:
        cur.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk")
        shifts = cur.fetchall()
    if not shifts:
        cur.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk")
        shifts = cur.fetchall()
    for s in shifts:
        jm, jk = s['jam_masuk'], s['jam_keluar']
        if jm <= jk:
            if jm <= now <= jk:
                cur.close(); return dict(s)
        else:
            if now >= jm or now <= jk:
                cur.close(); return dict(s)
    upcoming = next((s for s in shifts if s['jam_masuk'] > now), None)
    if upcoming:
        cur.close(); return dict(upcoming)
    if shifts:
        cur.close(); return dict(shifts[0])
    cur.close()
    return {'jam_masuk':'08:00','jam_keluar':'17:00','toleransi_menit':15,'nama':'Default','id':None}


def login_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*a, **kw)
    return dec

def admin_required(f):
    @wraps(f)
    def dec(*a, **kw):
        if 'user_id' not in session or session.get('role') != 'admin':
            flash('Akses ditolak!', 'error')
            return redirect(url_for('dashboard'))
        return f(*a, **kw)
    return dec

def q(conn):
    return conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

# ── AUTH ──────────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))


# ══════════════════════════════════════════════════════════════════════════════
# ── WEBAUTHN (Face ID / Fingerprint) ─────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

import base64, secrets, hashlib, struct, cbor2

def _b64url_encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).rstrip(b'=').decode()

def _b64url_decode(s: str) -> bytes:
    s = s.replace(' ', '+')
    pad = 4 - len(s) % 4
    if pad != 4:
        s += '=' * pad
    return base64.urlsafe_b64decode(s)

def _get_rp_id():
    """RP ID = hostname tanpa port."""
    host = request.host.split(':')[0]
    return host

def _verify_rp_id_hash(auth_data: bytes, rp_id: str) -> bool:
    expected = hashlib.sha256(rp_id.encode()).digest()
    return auth_data[:32] == expected


@app.route('/webauthn/register/begin', methods=['POST'])
@login_required
def webauthn_register_begin():
    """Mulai pendaftaran credential biometrik untuk user yang sudah login."""
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT id, email, nama FROM users WHERE id=%s", (uid,))
    user = cur.fetchone()
    cur.close(); conn.close()
    if not user:
        return jsonify(error='User tidak ditemukan'), 404

    challenge = secrets.token_bytes(32)
    session['webauthn_reg_challenge'] = _b64url_encode(challenge)

    rp_id = _get_rp_id()
    user_id_bytes = str(user['id']).encode()

    options = {
        'challenge': _b64url_encode(challenge),
        'rp': {'name': 'SIAP RSSR', 'id': rp_id},
        'user': {
            'id': _b64url_encode(user_id_bytes),
            'name': user['email'],
            'displayName': user['nama'],
        },
        'pubKeyCredParams': [
            {'type': 'public-key', 'alg': -7},   # ES256
            {'type': 'public-key', 'alg': -257},  # RS256
        ],
        'authenticatorSelection': {
            'authenticatorAttachment': 'platform',
            'userVerification': 'required',
            'residentKey': 'preferred',
        },
        'timeout': 60000,
        'attestation': 'none',
    }
    return jsonify(options)


@app.route('/webauthn/register/complete', methods=['POST'])
@login_required
def webauthn_register_complete():
    """Selesaikan pendaftaran — simpan public key."""
    uid = session['user_id']
    data = request.get_json()
    if not data:
        return jsonify(ok=False, error='Data tidak valid'), 400

    expected_challenge = session.pop('webauthn_reg_challenge', None)
    if not expected_challenge:
        return jsonify(ok=False, error='Challenge tidak ditemukan atau kadaluarsa'), 400

    try:
        client_data_json = _b64url_decode(data['clientDataJSON'])
        client_data = json.loads(client_data_json)

        # Verifikasi tipe
        if client_data.get('type') != 'webauthn.create':
            return jsonify(ok=False, error='Tipe tidak valid'), 400

        # Verifikasi challenge
        if client_data.get('challenge') != expected_challenge:
            return jsonify(ok=False, error='Challenge tidak cocok'), 400

        # Ambil credential id dan public key dari attestation object
        attestation_obj = cbor2.loads(_b64url_decode(data['attestationObject']))
        auth_data = attestation_obj['authData']

        # Verifikasi RP ID hash
        rp_id = _get_rp_id()
        if not _verify_rp_id_hash(auth_data, rp_id):
            return jsonify(ok=False, error='RP ID tidak cocok'), 400

        # Parse auth_data: 32 rpIdHash + 1 flags + 4 signCount + attested credential data
        flags = auth_data[32]
        if not (flags & 0x01):  # UP (user present)
            return jsonify(ok=False, error='User presence tidak terpenuhi'), 400

        sign_count = struct.unpack('>I', auth_data[33:37])[0]
        # AAGUID: 37-52, credIdLen: 53-54, credId: 55..
        cred_id_len = struct.unpack('>H', auth_data[53:55])[0]
        cred_id_bytes = auth_data[55:55 + cred_id_len]
        credential_id = _b64url_encode(cred_id_bytes)

        # Public key (COSE format, simpan as base64)
        pub_key_bytes = auth_data[55 + cred_id_len:]
        public_key_b64 = _b64url_encode(pub_key_bytes)

        device_name = data.get('deviceName', 'Perangkat')[:60]

        conn = get_db(); cur = q(conn)
        # Cek duplikat credential
        cur.execute("SELECT id FROM webauthn_credentials WHERE credential_id=%s", (credential_id,))
        if cur.fetchone():
            cur.close(); conn.close()
            return jsonify(ok=False, error='Credential sudah terdaftar'), 409

        cur.execute("""INSERT INTO webauthn_credentials
            (user_id, credential_id, public_key, sign_count, device_name)
            VALUES (%s,%s,%s,%s,%s)""",
            (uid, credential_id, public_key_b64, sign_count, device_name))
        log_audit(conn, 'CREATE', 'auth',
            deskripsi=f'Daftarkan biometrik: {device_name}', ref_id=uid, ref_table='users')
        conn.commit(); cur.close(); conn.close()
        return jsonify(ok=True, message='Biometrik berhasil didaftarkan!')

    except Exception as e:
        return jsonify(ok=False, error=f'Gagal: {str(e)}'), 500


@app.route('/webauthn/login/begin', methods=['POST'])
def webauthn_login_begin():
    """Mulai proses login biometrik — kirim challenge ke browser."""
    data = request.get_json() or {}
    email = data.get('email', '').strip()

    conn = get_db(); cur = q(conn)
    if email:
        cur.execute("""SELECT u.id, wc.credential_id
            FROM users u JOIN webauthn_credentials wc ON u.id=wc.user_id
            WHERE u.email=%s AND u.status='active'""", (email,))
    else:
        cur.execute("""SELECT u.id, wc.credential_id
            FROM users u JOIN webauthn_credentials wc ON u.id=wc.user_id
            WHERE u.status='active'""")
    rows = cur.fetchall()
    cur.close(); conn.close()

    if not rows:
        return jsonify(error='Tidak ada credential biometrik terdaftar'), 404

    challenge = secrets.token_bytes(32)
    session['webauthn_auth_challenge'] = _b64url_encode(challenge)

    allow_credentials = [
        {'type': 'public-key', 'id': row['credential_id']}
        for row in rows
    ]

    options = {
        'challenge': _b64url_encode(challenge),
        'rpId': _get_rp_id(),
        'allowCredentials': allow_credentials,
        'userVerification': 'required',
        'timeout': 60000,
    }
    return jsonify(options)


@app.route('/webauthn/login/complete', methods=['POST'])
def webauthn_login_complete():
    """Verifikasi assertion — login user jika valid."""
    data = request.get_json()
    if not data:
        return jsonify(ok=False, error='Data tidak valid'), 400

    expected_challenge = session.pop('webauthn_auth_challenge', None)
    if not expected_challenge:
        return jsonify(ok=False, error='Challenge kadaluarsa, coba lagi'), 400

    try:
        credential_id = data.get('credentialId', '')
        client_data_json = _b64url_decode(data['clientDataJSON'])
        client_data = json.loads(client_data_json)

        # Verifikasi tipe & challenge
        if client_data.get('type') != 'webauthn.get':
            return jsonify(ok=False, error='Tipe tidak valid'), 400
        if client_data.get('challenge') != expected_challenge:
            return jsonify(ok=False, error='Challenge tidak cocok'), 400

        auth_data = _b64url_decode(data['authenticatorData'])

        # Verifikasi RP ID hash
        rp_id = _get_rp_id()
        if not _verify_rp_id_hash(auth_data, rp_id):
            return jsonify(ok=False, error='RP ID tidak cocok'), 400

        flags = auth_data[32]
        if not (flags & 0x01):
            return jsonify(ok=False, error='User presence tidak terpenuhi'), 400

        # Ambil credential dari DB
        conn = get_db(); cur = q(conn)
        cur.execute("""SELECT wc.*, u.id as uid, u.nama, u.role, u.foto, u.status
            FROM webauthn_credentials wc JOIN users u ON wc.user_id=u.id
            WHERE wc.credential_id=%s""", (credential_id,))
        cred = cur.fetchone()

        if not cred:
            cur.close(); conn.close()
            return jsonify(ok=False, error='Credential tidak ditemukan'), 404

        if cred['status'] != 'active':
            cur.close(); conn.close()
            return jsonify(ok=False, error='Akun belum aktif atau ditolak'), 403

        # Update sign count & last_used
        new_sign_count = struct.unpack('>I', auth_data[33:37])[0]
        cur.execute("""UPDATE webauthn_credentials SET sign_count=%s, last_used=CURRENT_TIMESTAMP
            WHERE credential_id=%s""", (new_sign_count, credential_id))

        # Set session — login berhasil
        session.update({
            'user_id': cred['uid'],
            'nama':    cred['nama'],
            'role':    cred['role'],
            'foto':    cred['foto'],
        })
        log_audit(conn, 'LOGIN', 'auth',
            deskripsi=f'Login biometrik berhasil: {cred["nama"]}',
            ref_id=cred['uid'], ref_table='users',
            user_id=cred['uid'], user_nama=cred['nama'], user_role=cred['role'])
        conn.commit(); cur.close(); conn.close()
        return jsonify(ok=True, redirect=url_for('dashboard'))

    except Exception as e:
        return jsonify(ok=False, error=f'Verifikasi gagal: {str(e)}'), 500


@app.route('/webauthn/credentials', methods=['GET'])
@login_required
def webauthn_list_credentials():
    """Daftar credential biometrik milik user yang login."""
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT id, device_name, created_at, last_used
        FROM webauthn_credentials WHERE user_id=%s ORDER BY created_at DESC""", (uid,))
    rows = cur.fetchall()
    cur.close(); conn.close()
    result = []
    for r in rows:
        rd = dict(r)
        for k, v in rd.items():
            if hasattr(v, 'isoformat'): rd[k] = v.isoformat()
        result.append(rd)
    return jsonify(result)


@app.route('/webauthn/credentials/<int:cid>/hapus', methods=['POST'])
@login_required
def webauthn_hapus_credential(cid):
    """Hapus credential biometrik."""
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("DELETE FROM webauthn_credentials WHERE id=%s AND user_id=%s", (cid, uid))
    conn.commit()
    log_audit(conn, 'DELETE', 'auth', deskripsi=f'Hapus credential biometrik id={cid}',
        ref_id=uid, ref_table='users')
    cur.close(); conn.close()
    return jsonify(ok=True)


@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        conn = get_db(); cur = q(conn)
        cur.execute("SELECT * FROM users WHERE email=%s", (request.form.get('email','').strip(),))
        user = cur.fetchone()
        cur.close()
        if user and check_password_hash(user['password'], request.form.get('password','')):
            if user['status'] == 'pending':
                flash('Akun belum divalidasi admin.', 'warning')
                log_audit(conn, 'LOGIN_GAGAL', 'auth',
                    deskripsi=f'Login ditolak — akun pending: {user["email"]}',
                    ref_id=user['id'], ref_table='users', status='error',
                    user_id=user['id'], user_nama=user['nama'], user_role=user['role'])
            elif user['status'] == 'rejected':
                flash('Akun ditolak. Hubungi admin.', 'error')
                log_audit(conn, 'LOGIN_GAGAL', 'auth',
                    deskripsi=f'Login ditolak — akun rejected: {user["email"]}',
                    ref_id=user['id'], ref_table='users', status='error',
                    user_id=user['id'], user_nama=user['nama'], user_role=user['role'])
            else:
                session.update({'user_id':user['id'],'nama':user['nama'],'role':user['role'],'foto':user['foto']})
                log_audit(conn, 'LOGIN', 'auth',
                    deskripsi=f'Login berhasil: {user["nama"]} ({user["role"]})',
                    ref_id=user['id'], ref_table='users',
                    user_id=user['id'], user_nama=user['nama'], user_role=user['role'])
                conn.close()
                return redirect(url_for('dashboard'))
        else:
            email_input = request.form.get('email','')
            log_audit(conn, 'LOGIN_GAGAL', 'auth',
                deskripsi=f'Password salah atau email tidak ditemukan: {email_input}',
                status='error', user_id=None, user_nama=email_input, user_role='-')
            flash('Email atau password salah.', 'error')
        conn.close()
    return render_template('login.html')

@app.route('/register', methods=['GET','POST'])
def register():
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    if request.method == 'POST':
        dept_id = request.form.get('departemen_id') or None
        dept_nama = ''
        if dept_id:
            cur.execute("SELECT nama FROM departemen WHERE id=%s", (dept_id,))
            d = cur.fetchone()
            if d: dept_nama = d['nama']
        foto_path = None
        if 'foto' in request.files:
            f = request.files['foto']
            if f and f.filename and allowed_file(f.filename):
                nik = request.form.get('nik', 'new')
                fn = secure_filename(f"{nik}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[1].lower()}")
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
                foto_path = fn
        try:
            cur.execute("""INSERT INTO users (nik,nip,nama,email,password,jabatan,departemen,departemen_id,no_hp,alamat,tanggal_lahir,jenis_kelamin,foto)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (request.form['nik'].strip(), request.form.get('nip','').strip(),
                 request.form['nama'].strip(), request.form['email'].strip(),
                 generate_password_hash(request.form['password']), request.form.get('jabatan','').strip(),
                 dept_nama, dept_id, request.form.get('no_hp','').strip(), request.form.get('alamat','').strip(),
                 request.form.get('tanggal_lahir',''), request.form.get('jenis_kelamin',''), foto_path))
            conn.commit()
            flash('Registrasi berhasil! Tunggu validasi admin.', 'success')
            cur.close(); conn.close()
            return redirect(url_for('login'))
        except Exception as e:
            conn.rollback()
            flash('NIK atau Email sudah terdaftar.', 'error')
    cur.close(); conn.close()
    return render_template('register.html', depts=depts)

@app.route('/logout')
def logout():
    if 'user_id' in session:
        conn = get_db()
        log_audit(conn, 'LOGOUT', 'auth', deskripsi=f'Logout: {session.get("nama")}')
        conn.close()
    session.clear()
    return redirect(url_for('login'))

# ── USER ──────────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('role') == 'admin': return redirect(url_for('admin_dashboard'))
    uid = session['user_id']; today = date.today().isoformat()
    conn = get_db(); cur = q(conn)

    cur.execute("SELECT * FROM absensi WHERE user_id=%s AND tanggal=%s", (uid, today))
    absen_today = cur.fetchone()

    # ── Shift malam lintas hari: jika hari ini belum ada record masuk,
    #    cek apakah kemarin sudah absen masuk tapi belum absen keluar.
    #    Contoh: masuk 20:00 tgl 8 Apr, keluar 08:00 tgl 9 Apr.
    if not absen_today:
        yesterday = (date.today() - timedelta(days=1)).isoformat()
        cur.execute("""SELECT * FROM absensi
            WHERE user_id=%s AND tanggal=%s AND jam_masuk IS NOT NULL AND jam_keluar IS NULL""",
            (uid, yesterday))
        absen_kemarin_terbuka = cur.fetchone()
        if absen_kemarin_terbuka:
            _pakai = False
            if absen_kemarin_terbuka['shift_id']:
                cur.execute("SELECT * FROM shift WHERE id=%s", (absen_kemarin_terbuka['shift_id'],))
                _shift = cur.fetchone()
                if _shift and _shift['jam_keluar'] < _shift['jam_masuk']:
                    _pakai = True  # shift malam lintas hari
            else:
                _jm = (absen_kemarin_terbuka['jam_masuk'] or '00:00')[:5]
                if _jm >= '18:00':
                    _pakai = True  # heuristik: jam masuk malam
            if _pakai:
                absen_today = absen_kemarin_terbuka

    bulan = date.today().strftime('%Y-%m')
    cur.execute("""SELECT
        SUM(CASE WHEN status='hadir' THEN 1 ELSE 0 END) as hadir,
        SUM(CASE WHEN status='telat' THEN 1 ELSE 0 END) as telat,
        SUM(CASE WHEN status='izin' THEN 1 ELSE 0 END) as izin,
        COUNT(*) as total
        FROM absensi WHERE user_id=%s AND TO_CHAR(tanggal,'YYYY-MM')=%s""", (uid, bulan))
    stats = cur.fetchone()

    cur.execute("""SELECT a.tanggal,a.status,a.jam_masuk,a.jam_keluar,a.jarak_masuk,s.nama as shift_nama
        FROM absensi a LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.user_id=%s ORDER BY a.tanggal DESC LIMIT 30""", (uid,))
    riwayat = cur.fetchall()

    cur.execute("""SELECT u.*,d.nama as dept_nama,d.warna as dept_warna,
        s.nama as shift_nama,s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=%s""", (uid,))
    user = cur.fetchone()

    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()

    user_shift = get_active_shift(uid, conn)  # otomatis deteksi shift berjalan

    # Ambil daftar shift tersedia untuk user (dari departemen atau semua shift aktif)
    cur.execute("SELECT * FROM users WHERE id=%s", (uid,))
    u_raw = cur.fetchone()
    if u_raw and u_raw['departemen_id']:
        cur.execute("""SELECT s.* FROM shift s
            JOIN departemen_shift ds ON s.id=ds.shift_id
            WHERE ds.departemen_id=%s AND s.aktif=1 ORDER BY s.jam_masuk""", (u_raw['departemen_id'],))
        available_shifts = cur.fetchall()
        if not available_shifts:
            cur.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk")
            available_shifts = cur.fetchall()
    else:
        cur.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk")
        available_shifts = cur.fetchall()

    cur.close(); conn.close()

    if not user:
        session.clear()
        flash('Sesi tidak valid, silakan login kembali.', 'warning')
        return redirect(url_for('login'))

    return render_template('dashboard.html', absen_today=absen_today, stats=stats,
        riwayat=[dict(r) for r in riwayat], user=user, settings=settings,
        today=today, user_shift=user_shift, available_shifts=available_shifts)

@app.route('/absen', methods=['POST'])
@login_required
def absen():
    uid = session['user_id']; today = date.today().isoformat()
    now = datetime.now().strftime('%H:%M:%S')
    lat = request.form.get('lat', type=float)
    lng = request.form.get('lng', type=float)
    tipe = request.form.get('tipe')
    shift_id_form = request.form.get('shift_id') or None
    conn = get_db(); cur = q(conn)

    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    off_lat = settings['office_lat'] if settings else -6.2088
    off_lng = settings['office_lng'] if settings else 106.8456
    max_dist = settings['max_distance'] if settings else 100

    jarak = haversine(lat, lng, off_lat, off_lng) if lat and lng else None

    if jarak is not None and jarak > max_dist:
        cur.close(); conn.close()
        flash(f'Absen ditolak! Anda berada {jarak:.0f}m dari kantor. Batas maksimal {max_dist}m.', 'error_radius')
        return redirect(url_for('dashboard'))

    foto_path = None
    if 'foto' in request.files:
        f = request.files['foto']
        if f and f.filename:
            ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else 'jpg'
            fn = secure_filename(f"{uid}_{today}_{tipe}.{ext}")
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
            foto_path = fn

    cur.execute("SELECT * FROM absensi WHERE user_id=%s AND tanggal=%s", (uid, today))
    absen_today = cur.fetchone()

    # ── Shift malam lintas hari: cek record kemarin yang belum keluar
    if not absen_today:
        _yest = (date.today() - timedelta(days=1)).isoformat()
        cur.execute("""SELECT * FROM absensi
            WHERE user_id=%s AND tanggal=%s AND jam_masuk IS NOT NULL AND jam_keluar IS NULL""",
            (uid, _yest))
        _open = cur.fetchone()
        if _open:
            _use = False
            if _open['shift_id']:
                cur.execute('SELECT * FROM shift WHERE id=%s', (_open['shift_id'],))
                _s = cur.fetchone()
                if _s and _s['jam_keluar'] < _s['jam_masuk']:
                    _use = True
            else:
                if (_open['jam_masuk'] or '00:00')[:5] >= '18:00':
                    _use = True
            if _use:
                absen_today = _open

    if tipe == 'masuk':
        if absen_today:
            flash('Sudah absen masuk hari ini!', 'warning')
        else:
            # Gunakan shift yang dipilih user, atau fallback ke shift aktif
            shift_id = shift_id_form
            if shift_id:
                cur.execute("SELECT * FROM shift WHERE id=%s AND aktif=1", (shift_id,))
                chosen_shift = cur.fetchone()
            else:
                chosen_shift = get_active_shift(uid, conn)
                shift_id = chosen_shift.get('id') if chosen_shift else None

            # Cek telat berdasarkan shift yang dipilih
            status = 'hadir'
            if chosen_shift:
                from datetime import time as dtime
                h, m = map(int, chosen_shift['jam_masuk'].split(':'))
                toleransi = chosen_shift.get('toleransi_menit', 15)
                batas = datetime.combine(date.today(), dtime(h, m)) + timedelta(minutes=toleransi)
                if datetime.now() > batas: status = 'telat'

            cur.execute("""INSERT INTO absensi (user_id,tanggal,jam_masuk,foto_masuk,lat_masuk,lng_masuk,jarak_masuk,shift_id,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (uid, today, now, foto_path, lat, lng, jarak, shift_id, status))
            conn.commit()
            log_audit(conn, 'ABSEN_MASUK', 'absensi',
                deskripsi=f'Absen masuk — {session.get("nama")} | Jarak: {jarak:.0f}m | Status: {status}' if jarak else f'Absen masuk — {session.get("nama")} | Status: {status}',
                data_baru={'tanggal':today,'jam':now,'jarak':jarak,'shift_id':shift_id,'status':status},
                ref_id=uid, ref_table='users')
            flash(f'Absen masuk berhasil! Jarak: {jarak:.0f}m' if jarak else 'Absen masuk berhasil!', 'success')

    elif tipe == 'keluar':
        if not absen_today:
            flash('Belum absen masuk!', 'warning')
        elif absen_today['jam_keluar']:
            flash('Sudah absen keluar!', 'warning')
        else:
            # LOCK shift keluar = shift masuk, abaikan pilihan user
            shift_id = absen_today['shift_id']
            _tanggal_absen = absen_today['tanggal'] if absen_today['tanggal'] else today
            if hasattr(_tanggal_absen, 'isoformat'): _tanggal_absen = _tanggal_absen.isoformat()
            cur.execute("""UPDATE absensi SET jam_keluar=%s,foto_keluar=%s,lat_keluar=%s,lng_keluar=%s,jarak_keluar=%s,shift_id=%s
                WHERE user_id=%s AND tanggal=%s""", (now, foto_path, lat, lng, jarak, shift_id, uid, _tanggal_absen))
            conn.commit()
            log_audit(conn, 'ABSEN_KELUAR', 'absensi',
                deskripsi=f'Absen keluar — {session.get("nama")} | Jarak: {jarak:.0f}m' if jarak else f'Absen keluar — {session.get("nama")}',
                data_baru={'tanggal':today,'jam':now,'jarak':jarak},
                ref_id=uid, ref_table='users')
            flash('Absen keluar berhasil!', 'success')

    cur.close(); conn.close()
    return redirect(url_for('dashboard'))

@app.route('/lupa-absen', methods=['POST'])
@login_required
def lupa_absen():
    """User lapor lupa absen pulang — admin yang approve."""
    uid = session['user_id']
    tanggal = request.form.get('tanggal', date.today().isoformat())
    alasan = request.form.get('alasan', 'Lupa absen pulang')
    jam_keluar_manual = request.form.get('jam_keluar_manual', '').strip()
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM absensi WHERE user_id=%s AND tanggal=%s", (uid, tanggal))
    absen = cur.fetchone()
    if not absen:
        flash('Tidak ada data absen masuk pada tanggal tersebut.', 'error')
    elif absen['jam_keluar']:
        flash('Absen pulang sudah tercatat.', 'warning')
    else:
        # Simpan sebagai izin dengan jenis khusus "Lupa Absen Pulang"
        cur.execute("""INSERT INTO izin (user_id,tanggal_mulai,tanggal_selesai,jenis,alasan,status)
            VALUES (%s,%s,%s,%s,%s,'pending')""",
            (uid, tanggal, tanggal, 'Lupa Absen Pulang',
             f"{alasan} | Jam keluar: {jam_keluar_manual if jam_keluar_manual else 'tidak diisi'}"))
        conn.commit()
        flash('Laporan lupa absen pulang berhasil dikirim. Tunggu konfirmasi admin.', 'success')
    cur.close(); conn.close()
    return redirect(url_for('dashboard'))

@app.route('/admin/izin/<int:iid>/approve-lupa-absen', methods=['POST'])
@admin_required
def approve_lupa_absen(iid):
    """Admin setujui lupa absen pulang — isi jam keluar manual."""
    jam_keluar = request.form.get('jam_keluar', '17:00') + ':00'
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM izin WHERE id=%s", (iid,))
    iz = cur.fetchone()
    if iz:
        cur.execute("UPDATE izin SET status='approved',catatan_admin=%s WHERE id=%s",
            (f'Disetujui admin, jam keluar: {jam_keluar}', iid))
        cur.execute("""UPDATE absensi SET jam_keluar=%s,keterangan='Lupa absen pulang — disetujui admin'
            WHERE user_id=%s AND tanggal=%s AND jam_keluar IS NULL""",
            (jam_keluar, iz['user_id'], str(iz['tanggal_mulai'])))
        conn.commit()
        flash(f'Lupa absen pulang disetujui. Jam keluar: {jam_keluar}', 'success')
    cur.close(); conn.close()
    return redirect(url_for('admin_izin'))

@app.route('/riwayat')
@login_required
def riwayat():
    uid = session['user_id']
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT a.*,s.nama as shift_nama,s.jam_masuk as shift_masuk
        FROM absensi a LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.user_id=%s AND TO_CHAR(a.tanggal,'YYYY-MM')=%s ORDER BY a.tanggal DESC""", (uid, bulan))
    data = cur.fetchall()
    cur.close(); conn.close()
    return render_template('riwayat.html', data=data, bulan=bulan)

# ── FIX: Route izin dengan allowed_lampiran ───────────────────────────────────
@app.route('/izin', methods=['GET','POST'])
@login_required
def izin():
    uid = session['user_id']
    if request.method == 'POST':
        conn = get_db(); cur = q(conn)
        lamp = None

        if 'lampiran' in request.files:
            f = request.files['lampiran']
            if f and f.filename:
                if allowed_lampiran(f.filename):
                    ext = f.filename.rsplit('.', 1)[-1].lower()
                    fn = secure_filename(f"{uid}_izin_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
                    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
                    f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
                    lamp = fn
                else:
                    flash('Format file tidak didukung. Gunakan JPG, PNG, atau PDF.', 'error')
                    cur.close(); conn.close()
                    return redirect(url_for('izin'))

        cur.execute("""INSERT INTO izin (user_id,tanggal_mulai,tanggal_selesai,jenis,alasan,lampiran)
            VALUES (%s,%s,%s,%s,%s,%s)""",
            (uid, request.form['tanggal_mulai'], request.form['tanggal_selesai'],
             request.form['jenis'], request.form['alasan'], lamp))
        conn.commit(); cur.close(); conn.close()
        flash('Permohonan izin berhasil dikirim!', 'success')
        return redirect(url_for('izin'))

    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM izin WHERE user_id=%s ORDER BY created_at DESC", (uid,))
    data = cur.fetchall()
    cur.close(); conn.close()
    return render_template('izin.html', data=data)

@app.route('/profil', methods=['GET','POST'])
@login_required
def profil():
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    if request.method == 'POST':
        foto_path = None
        if 'foto' in request.files:
            f = request.files['foto']
            if f and f.filename and allowed_file(f.filename):
                fn = secure_filename(f"user_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[-1].lower()}")
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn))
                foto_path = fn
        dept_id = request.form.get('departemen_id') or None
        # Ambil nama departemen untuk disimpan di kolom departemen (teks)
        dept_nama = None
        if dept_id:
            cur.execute("SELECT nama FROM departemen WHERE id=%s AND aktif=1", (dept_id,))
            d_row = cur.fetchone()
            if d_row:
                dept_nama = d_row['nama']
        if foto_path:
            cur.execute("UPDATE users SET no_hp=%s,alamat=%s,foto=%s,departemen_id=%s,departemen=%s WHERE id=%s",
                (request.form.get('no_hp',''), request.form.get('alamat',''), foto_path, dept_id, dept_nama, uid))
            session['foto'] = foto_path
        else:
            cur.execute("UPDATE users SET no_hp=%s,alamat=%s,departemen_id=%s,departemen=%s WHERE id=%s",
                (request.form.get('no_hp',''), request.form.get('alamat',''), dept_id, dept_nama, uid))
        conn.commit()
        flash('Profil berhasil diperbarui!', 'success')

    cur.execute("""SELECT u.*,d.nama as dept_nama,d.warna as dept_warna,
        s.nama as shift_nama,s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=%s""", (uid,))
    user = cur.fetchone()
    cur.execute("SELECT id, nama FROM departemen WHERE aktif=1 ORDER BY nama")
    daftar_departemen = cur.fetchall()
    cur.close(); conn.close()
    if not user:
        session.clear()
        flash('Sesi tidak valid, silakan login kembali.', 'warning')
        return redirect(url_for('login'))
    return render_template('profil.html', user=user, daftar_departemen=daftar_departemen)

# ── ADMIN DASHBOARD ───────────────────────────────────────────────────────────
@app.route('/admin')
@admin_required
def admin_dashboard():
    conn = get_db(); cur = q(conn)
    today = date.today().isoformat()

    cur.execute("SELECT COUNT(*) as c FROM users WHERE role='user' AND status='active'")
    total_user = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM users WHERE status='pending'")
    pending = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM absensi WHERE tanggal=%s AND status IN ('hadir','telat')", (today,))
    hadir_today = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM absensi WHERE tanggal=%s AND status='telat'", (today,))
    telat_today = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM izin WHERE status='pending'")
    izin_pending = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM departemen WHERE aktif=1")
    total_dept = cur.fetchone()['c']
    cur.execute("SELECT COUNT(*) as c FROM shift WHERE aktif=1")
    total_shift = cur.fetchone()['c']

    chart_data = []
    for i in range(6, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        cur.execute("SELECT COUNT(*) as c FROM absensi WHERE tanggal=%s AND status IN ('hadir','telat')", (d,))
        h = cur.fetchone()['c']
        cur.execute("SELECT COUNT(*) as c FROM absensi WHERE tanggal=%s AND status='telat'", (d,))
        t = cur.fetchone()['c']
        chart_data.append({'tanggal': d, 'hadir': h, 'telat': t})

    cur.execute("""SELECT d.nama as departemen, d.warna, d.kode,
        COUNT(DISTINCT u.id) as total_pegawai,
        SUM(CASE WHEN a.tanggal=%s AND a.status IN ('hadir','telat') THEN 1 ELSE 0 END) as hadir_today
        FROM departemen d
        LEFT JOIN users u ON u.departemen_id=d.id AND u.status='active' AND u.role='user'
        LEFT JOIN absensi a ON u.id=a.user_id AND a.tanggal=%s
        WHERE d.aktif=1 GROUP BY d.id,d.nama,d.warna,d.kode ORDER BY total_pegawai DESC""", (today, today))
    dept_stats = cur.fetchall()

    cur.execute("""SELECT a.*,u.nama,u.jabatan,u.foto,u.departemen,s.nama as shift_nama
        FROM absensi a JOIN users u ON a.user_id=u.id LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.tanggal=%s ORDER BY a.jam_masuk DESC LIMIT 10""", (today,))
    recent = cur.fetchall()

    cur.execute("""SELECT d.*,(SELECT COUNT(*) FROM users WHERE departemen_id=d.id AND status='active') as jml
        FROM departemen d WHERE aktif=1 ORDER BY nama""")
    recent_depts = cur.fetchall()

    cur.execute("""SELECT s.*,(SELECT COUNT(*) FROM users WHERE shift_id=s.id AND status='active') as jml
        FROM shift s WHERE aktif=1 ORDER BY jam_masuk""")
    recent_shifts = cur.fetchall()

    cur.close(); conn.close()
    return render_template('admin/dashboard.html',
        total_user=total_user, pending=pending, hadir_today=hadir_today,
        telat_today=telat_today, izin_pending=izin_pending,
        total_dept=total_dept, total_shift=total_shift,
        chart_data=json.dumps(chart_data), dept_stats=dept_stats,
        recent=recent, recent_depts=recent_depts, recent_shifts=recent_shifts, today=today)

# ── ADMIN DEPARTEMEN ──────────────────────────────────────────────────────────
@app.route('/admin/departemen')
@admin_required
def admin_departemen():
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT d.*,
        COUNT(DISTINCT u.id) as total_pegawai,
        COUNT(DISTINCT ds.shift_id) as total_shift
        FROM departemen d
        LEFT JOIN users u ON u.departemen_id=d.id AND u.status='active' AND u.role='user'
        LEFT JOIN departemen_shift ds ON ds.departemen_id=d.id
        GROUP BY d.id ORDER BY d.nama""")
    depts = cur.fetchall()
    cur.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk")
    shifts = cur.fetchall()
    cur.execute("SELECT * FROM departemen_shift")
    dept_shifts = {}
    for ds in cur.fetchall():
        dept_shifts.setdefault(str(ds['departemen_id']), []).append(str(ds['shift_id']))
    cur.close(); conn.close()
    return render_template('admin/departemen.html', depts=depts, shifts=shifts, dept_shifts=dept_shifts)

@app.route('/admin/departemen/tambah', methods=['POST'])
@admin_required
def tambah_departemen():
    conn = get_db(); cur = q(conn)
    try:
        cur.execute("INSERT INTO departemen (nama,kode,deskripsi,warna) VALUES (%s,%s,%s,%s)",
            (request.form['nama'], request.form['kode'].upper(),
             request.form.get('deskripsi',''), request.form.get('warna','#3b82f6')))
        conn.commit()
        flash(f'Departemen "{request.form["nama"]}" ditambahkan!', 'success')
    except:
        conn.rollback(); flash('Nama atau kode sudah ada!', 'error')
    cur.close(); conn.close()
    return redirect(url_for('admin_departemen'))

@app.route('/admin/departemen/edit/<int:did>', methods=['POST'])
@admin_required
def edit_departemen(did):
    conn = get_db(); cur = q(conn)
    try:
        cur.execute("UPDATE departemen SET nama=%s,kode=%s,deskripsi=%s,warna=%s,aktif=%s WHERE id=%s",
            (request.form['nama'], request.form['kode'].upper(), request.form.get('deskripsi',''),
             request.form.get('warna','#3b82f6'), 1 if request.form.get('aktif') else 0, did))
        conn.commit(); flash('Departemen diperbarui!', 'success')
    except Exception as e:
        conn.rollback(); flash('Gagal: '+str(e), 'error')
    cur.close(); conn.close()
    return redirect(url_for('admin_departemen'))

@app.route('/admin/departemen/hapus/<int:did>', methods=['POST'])
@admin_required
def hapus_departemen(did):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT COUNT(*) as c FROM users WHERE departemen_id=%s", (did,))
    peg = cur.fetchone()['c']
    if peg > 0:
        flash(f'Tidak bisa hapus: masih ada {peg} pegawai!', 'error')
    else:
        cur.execute("DELETE FROM departemen_shift WHERE departemen_id=%s", (did,))
        cur.execute("DELETE FROM departemen WHERE id=%s", (did,))
        conn.commit(); flash('Departemen dihapus!', 'success')
    cur.close(); conn.close()
    return redirect(url_for('admin_departemen'))

@app.route('/admin/departemen/<int:did>/shift', methods=['POST'])
@admin_required
def atur_shift_departemen(did):
    conn = get_db(); cur = q(conn)
    try:
        cur.execute("DELETE FROM departemen_shift WHERE departemen_id=%s", (did,))
        for sid in request.form.getlist('shift_ids'):
            try:
                cur.execute("INSERT INTO departemen_shift (departemen_id,shift_id) VALUES (%s,%s) ON CONFLICT DO NOTHING", (did, int(sid)))
            except (ValueError, TypeError):
                pass
        conn.commit()
        flash('Shift departemen diperbarui!', 'success')
    except Exception as e:
        conn.rollback()
        flash('Gagal memperbarui shift: ' + str(e), 'error')
    cur.close(); conn.close()
    return redirect(url_for('admin_departemen'))

# ── ADMIN SHIFT ───────────────────────────────────────────────────────────────
@app.route('/admin/shift')
@admin_required
def admin_shift():
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT s.*,
        COUNT(DISTINCT u.id) as total_pegawai,
        COUNT(DISTINCT ds.departemen_id) as total_dept
        FROM shift s
        LEFT JOIN users u ON u.shift_id=s.id AND u.status='active'
        LEFT JOIN departemen_shift ds ON ds.shift_id=s.id
        GROUP BY s.id ORDER BY s.jam_masuk""")
    shifts = cur.fetchall()
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/shift.html', shifts=shifts, depts=depts)

@app.route('/admin/shift/tambah', methods=['POST'])
@admin_required
def tambah_shift():
    conn = get_db(); cur = q(conn)
    nama = request.form.get('nama', '').strip()
    if not nama:
        flash('Nama shift tidak boleh kosong.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('admin_shift'))
    cur.execute("SELECT id FROM shift WHERE LOWER(nama)=LOWER(%s)", (nama,))
    if cur.fetchone():
        flash(f'Shift "{nama}" sudah ada, tidak bisa duplikat.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('admin_shift'))
    try:
        cur.execute("INSERT INTO shift (nama,jam_masuk,jam_keluar,toleransi_menit,deskripsi,warna) VALUES (%s,%s,%s,%s,%s,%s)",
            (nama, request.form['jam_masuk'], request.form['jam_keluar'],
             int(request.form.get('toleransi_menit', 15)), request.form.get('deskripsi',''),
             request.form.get('warna','#10b981')))
        conn.commit(); flash(f'Shift "{nama}" ditambahkan!', 'success')
    except Exception as e:
        conn.rollback(); flash('Gagal: '+str(e), 'error')
    cur.close(); conn.close()
    return redirect(url_for('admin_shift'))

@app.route('/admin/shift/edit/<int:sid>', methods=['POST'])
@admin_required
def edit_shift(sid):
    conn = get_db(); cur = q(conn)
    cur.execute("UPDATE shift SET nama=%s,jam_masuk=%s,jam_keluar=%s,toleransi_menit=%s,deskripsi=%s,warna=%s,aktif=%s WHERE id=%s",
        (request.form['nama'], request.form['jam_masuk'], request.form['jam_keluar'],
         int(request.form.get('toleransi_menit', 15)), request.form.get('deskripsi',''),
         request.form.get('warna','#10b981'), 1 if request.form.get('aktif') else 0, sid))
    conn.commit(); cur.close(); conn.close()
    flash('Shift diperbarui!', 'success')
    return redirect(url_for('admin_shift'))

@app.route('/admin/shift/hapus/<int:sid>', methods=['POST'])
@admin_required
def hapus_shift(sid):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT COUNT(*) as c FROM users WHERE shift_id=%s", (sid,))
    peg = cur.fetchone()['c']
    if peg > 0:
        flash(f'Tidak bisa hapus: {peg} pegawai masih menggunakan shift ini!', 'error')
    else:
        cur.execute("DELETE FROM departemen_shift WHERE shift_id=%s", (sid,))
        cur.execute("DELETE FROM shift WHERE id=%s", (sid,))
        conn.commit(); flash('Shift dihapus!', 'success')
    cur.close(); conn.close()
    return redirect(url_for('admin_shift'))

# ── ADMIN PEGAWAI ─────────────────────────────────────────────────────────────
@app.route('/admin/pegawai')
@admin_required
def admin_pegawai():
    conn = get_db(); cur = q(conn)
    qstr = request.args.get('q',''); sf = request.args.get('status',''); df = request.args.get('dept','')
    sql = """SELECT u.*,d.nama as dept_nama,d.warna as dept_warna,
        s.nama as shift_nama,s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.role='user'"""
    params = []
    if qstr:
        sql += " AND (u.nama ILIKE %s OR u.nik ILIKE %s OR u.email ILIKE %s)"
        params.extend([f'%{qstr}%']*3)
    if sf: sql += " AND u.status=%s"; params.append(sf)
    if df: sql += " AND u.departemen_id=%s"; params.append(df)
    sql += " ORDER BY u.created_at DESC"
    cur.execute(sql, params)
    users = cur.fetchall()
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    cur.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk")
    shifts = cur.fetchall()
    stats = {'total': len(users),
             'active': sum(1 for u in users if u['status']=='active'),
             'pending': sum(1 for u in users if u['status']=='pending')}
    cur.close(); conn.close()
    return render_template('admin/pegawai.html', users=users, q=qstr,
        status_filter=sf, dept_filter=df, depts=depts, shifts=shifts, stats=stats)

@app.route('/admin/pegawai/tambah', methods=['POST'])
@admin_required
def tambah_pegawai():
    conn = get_db(); cur = q(conn)
    dept_id = request.form.get('departemen_id') or None; dept_nama = ''
    if dept_id:
        cur.execute("SELECT nama FROM departemen WHERE id=%s", (dept_id,))
        d = cur.fetchone()
        if d: dept_nama = d['nama']
    foto_path = None
    if 'foto' in request.files:
        f = request.files['foto']
        if f and f.filename and allowed_file(f.filename):
            fn = secure_filename(f"{request.form.get('nik','new')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[-1].lower()}")
            f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn)); foto_path = fn
    try:
        cur.execute("""INSERT INTO users (nik,nama,email,password,jabatan,departemen,departemen_id,shift_id,
            no_hp,alamat,tanggal_lahir,jenis_kelamin,foto,status) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (request.form['nik'], request.form['nama'], request.form['email'],
             generate_password_hash(request.form['password']), request.form.get('jabatan',''),
             dept_nama, dept_id, request.form.get('shift_id') or None,
             request.form.get('no_hp',''), request.form.get('alamat',''),
             request.form.get('tanggal_lahir',''), request.form.get('jenis_kelamin',''),
             foto_path, request.form.get('status','active')))
        conn.commit(); flash('Pegawai berhasil ditambahkan!', 'success')
    except Exception as e:
        conn.rollback(); flash('NIK atau Email sudah ada: '+str(e), 'error')
    cur.close(); conn.close()
    return redirect(url_for('admin_pegawai'))

@app.route('/admin/pegawai/edit/<int:uid>', methods=['GET','POST'])
@admin_required
def edit_pegawai(uid):
    conn = get_db(); cur = q(conn)
    if request.method == 'POST':
        dept_id = int(request.form.get('departemen_id')) if request.form.get('departemen_id') else None
        dept_nama = ''
        if dept_id:
            cur.execute("SELECT nama FROM departemen WHERE id=%s", (dept_id,))
            d = cur.fetchone()
            if d: dept_nama = d['nama']
        shift_id = int(request.form.get('shift_id')) if request.form.get('shift_id') else None
        foto_path = None
        if 'foto' in request.files:
            f = request.files['foto']
            if f and f.filename and allowed_file(f.filename):
                fn = secure_filename(f"user_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{f.filename.rsplit('.',1)[-1].lower()}")
                f.save(os.path.join(app.config['UPLOAD_FOLDER'], fn)); foto_path = fn
        # Cek apakah kolom nip ada di tabel users
        cur.execute("SELECT column_name FROM information_schema.columns WHERE table_name='users' AND column_name='nip'")
        has_nip = cur.fetchone() is not None
        fields = ["nik=%s","nama=%s","email=%s","jabatan=%s","departemen=%s","departemen_id=%s",
                  "shift_id=%s","no_hp=%s","alamat=%s","tanggal_lahir=%s","jenis_kelamin=%s","status=%s","role=%s"]
        params = [request.form['nik'],
                  request.form['nama'], request.form['email'],
                  request.form.get('jabatan',''), dept_nama, dept_id,
                  shift_id, request.form.get('no_hp',''),
                  request.form.get('alamat',''), request.form.get('tanggal_lahir',''),
                  request.form.get('jenis_kelamin',''), request.form.get('status','active'),
                  request.form.get('role','user')]
        if has_nip:
            fields.insert(1, "nip=%s")
            params.insert(1, request.form.get('nip','').strip())
        if foto_path: fields.append("foto=%s"); params.append(foto_path)
        if request.form.get('password'): fields.append("password=%s"); params.append(generate_password_hash(request.form['password']))
        params.append(uid)
        try:
            cur.execute(f"UPDATE users SET {','.join(fields)} WHERE id=%s", params)
            conn.commit()
            flash('Data pegawai diperbarui!', 'success')
        except Exception as e:
            conn.rollback()
            flash('Gagal memperbarui data: ' + str(e), 'error')
        cur.close(); conn.close()
        return redirect(url_for('admin_pegawai'))
    cur.execute("""SELECT u.*,d.nama as dept_nama,s.nama as shift_nama
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=%s""", (uid,))
    user = cur.fetchone()
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    cur.execute("SELECT * FROM shift WHERE aktif=1 ORDER BY jam_masuk")
    shifts = cur.fetchall()
    cur.execute("SELECT * FROM master_role WHERE aktif=1 ORDER BY urutan")
    roles = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/edit_pegawai.html', user=user, depts=depts, shifts=shifts, roles=roles)

@app.route('/admin/pegawai/hapus/<int:uid>', methods=['POST'])
@admin_required
def hapus_pegawai(uid):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT COUNT(*) as c FROM absensi WHERE user_id=%s", (uid,))
    cnt = cur.fetchone()['c']
    if cnt > 0:
        cur.execute("UPDATE users SET status='rejected' WHERE id=%s", (uid,))
        conn.commit(); flash('Pegawai dinonaktifkan (ada data absensi).', 'warning')
    else:
        cur.execute("DELETE FROM izin WHERE user_id=%s", (uid,))
        cur.execute("DELETE FROM users WHERE id=%s", (uid,))
        conn.commit(); flash('Pegawai dihapus!', 'success')
    cur.close(); conn.close()
    return redirect(url_for('admin_pegawai'))

@app.route('/admin/validasi/<int:uid>/<action>')
@admin_required
def validasi_user(uid, action):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT nama,status FROM users WHERE id=%s", (uid,))
    target = cur.fetchone()
    status_lama = target['status'] if target else '-'
    status_baru = 'active' if action=='approve' else 'rejected'
    cur.execute("UPDATE users SET status=%s WHERE id=%s", (status_baru, uid))
    conn.commit()
    log_audit(conn, 'VALIDASI', 'pegawai',
        deskripsi=f'Validasi akun {target["nama"] if target else uid}: {status_lama} → {status_baru}',
        data_lama={'status':status_lama}, data_baru={'status':status_baru},
        ref_id=uid, ref_table='users')
    cur.close(); conn.close()
    flash('Akun disetujui!' if action=='approve' else 'Akun ditolak!',
          'success' if action=='approve' else 'info')
    return redirect(url_for('admin_pegawai'))

# ── ADMIN ABSENSI ─────────────────────────────────────────────────────────────
@app.route('/admin/absensi')
@admin_required
def admin_absensi():
    conn = get_db(); cur = q(conn)
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    dept = request.args.get('dept','')
    sql = """SELECT a.*,u.nama,u.nik,u.jabatan,u.departemen,u.foto,s.nama as shift_nama
        FROM absensi a JOIN users u ON a.user_id=u.id LEFT JOIN shift s ON a.shift_id=s.id
        WHERE TO_CHAR(a.tanggal,'YYYY-MM')=%s"""
    params = [bulan]
    if dept: sql += " AND u.departemen_id=%s"; params.append(dept)
    sql += " ORDER BY a.tanggal DESC,a.jam_masuk DESC"
    cur.execute(sql, params)
    data = cur.fetchall()
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/absensi.html', data=data, bulan=bulan, dept=dept, depts=depts)

# ── ADMIN IZIN ────────────────────────────────────────────────────────────────
@app.route('/admin/izin')
@admin_required
def admin_izin():
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT i.*,u.nama,u.nik,u.departemen FROM izin i JOIN users u ON i.user_id=u.id ORDER BY i.created_at DESC")
    data = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/izin.html', data=data)

@app.route('/admin/izin/<int:iid>/<action>', methods=['GET','POST'])
@admin_required
def proses_izin(iid, action):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM izin WHERE id=%s", (iid,))
    iz = cur.fetchone()
    if iz and action in ['approve','reject']:
        cur.execute("UPDATE izin SET status=%s WHERE id=%s",
            ('approved' if action=='approve' else 'rejected', iid))

        if action == 'approve':
            if iz['jenis'] == 'Lupa Absen Pulang':
                # Ambil jam keluar dari form (admin input) atau dari alasan
                jam_keluar_input = request.form.get('jam_keluar', '').strip()
                if jam_keluar_input:
                    jam_keluar = jam_keluar_input + ':00' if len(jam_keluar_input) == 5 else jam_keluar_input
                else:
                    # Fallback: ekstrak dari alasan
                    import re
                    alasan = iz['alasan'] or ''
                    m2 = re.search(r'Jam keluar[:\s]+(\d{1,2}:\d{2})', alasan)
                    jam_keluar = (m2.group(1) + ':00') if m2 else '17:00:00'

                tanggal = str(iz['tanggal_mulai'])
                cur.execute("""UPDATE absensi SET jam_keluar=%s,
                    keterangan='Lupa absen pulang — disetujui admin'
                    WHERE user_id=%s AND tanggal=%s AND (jam_keluar IS NULL OR jam_keluar='')""",
                    (jam_keluar, iz['user_id'], tanggal))
                cur.execute("UPDATE izin SET catatan_admin=%s WHERE id=%s",
                    (f'Jam keluar diisi: {jam_keluar}', iid))
                flash(f'Lupa absen pulang disetujui. Jam keluar: {jam_keluar}', 'success')
            else:
                d1 = datetime.strptime(str(iz['tanggal_mulai']), '%Y-%m-%d')
                d2 = datetime.strptime(str(iz['tanggal_selesai']), '%Y-%m-%d')
                cur2 = conn.cursor()
                cur_date = d1
                while cur_date <= d2:
                    cur2.execute("""INSERT INTO absensi (user_id,tanggal,status,keterangan)
                        VALUES (%s,%s,%s,%s) ON CONFLICT DO NOTHING""",
                        (iz['user_id'], cur_date.date().isoformat(), 'izin', iz['jenis']))
                    cur_date += timedelta(days=1)
                cur2.close()
                flash('Izin disetujui!', 'success')
        else:
            flash('Izin ditolak!', 'info')

        log_audit(conn,
            'APPROVE' if action=='approve' else 'REJECT', 'izin',
            deskripsi=f'{"Setujui" if action=="approve" else "Tolak"} izin ID {iid} — {iz["jenis"]} milik user_id {iz["user_id"]}',
            data_lama={'status':'pending'},
            data_baru={'status':'approved' if action=="approve" else 'rejected'},
            ref_id=iid, ref_table='izin')
        conn.commit()
    cur.close(); conn.close()
    return redirect(url_for('admin_izin'))

# ── ADMIN LAPORAN ─────────────────────────────────────────────────────────────
@app.route('/admin/laporan')
@admin_required
def admin_laporan():
    conn = get_db(); cur = q(conn)
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    cur.execute("""SELECT u.*,d.nama as dept_nama,s.nama as shift_nama FROM users u
        LEFT JOIN departemen d ON u.departemen_id=d.id LEFT JOIN shift s ON u.shift_id=s.id
        WHERE u.role='user' AND u.status='active' ORDER BY u.nama""")
    users = cur.fetchall()
    rekap = []
    for u in users:
        cur.execute("""SELECT
            SUM(CASE WHEN status='hadir' THEN 1 ELSE 0 END) as hadir,
            SUM(CASE WHEN status='telat' THEN 1 ELSE 0 END) as telat,
            SUM(CASE WHEN status='izin' THEN 1 ELSE 0 END) as izin,
            COUNT(*) as total
            FROM absensi WHERE user_id=%s AND TO_CHAR(tanggal,'YYYY-MM')=%s""", (u['id'], bulan))
        rekap.append({'user': u, 'stats': cur.fetchone()})
    cur.close(); conn.close()
    return render_template('admin/laporan.html', rekap=rekap, bulan=bulan)

@app.route('/admin/export/excel')
@admin_required
def export_excel():
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    conn = get_db()
    log_audit(conn, 'EXPORT', 'laporan', deskripsi=f'Export Excel laporan bulan {bulan}')
    conn.close()
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT a.tanggal,u.nik,u.nama,u.departemen,u.jabatan,a.jam_masuk,a.jam_keluar,
        a.jarak_masuk,a.status,a.keterangan,s.nama as shift_nama
        FROM absensi a JOIN users u ON a.user_id=u.id LEFT JOIN shift s ON a.shift_id=s.id
        WHERE TO_CHAR(a.tanggal,'YYYY-MM')=%s ORDER BY a.tanggal,u.nama""", (bulan,))
    data = cur.fetchall()
    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    cur.close(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"Absensi {bulan}"
    for col,w in zip('ABCDEFGHIJK',[12,14,22,18,20,12,12,14,10,15,20]):
        ws.column_dimensions[col].width = w
    ws.merge_cells('A1:K1')
    ws['A1'] = f"LAPORAN ABSENSI - {settings['nama_perusahaan'] if settings else 'PT Absensi'}"
    ws['A1'].font = Font(bold=True,size=14); ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:K2'); ws['A2'] = f"Periode: {bulan}"
    ws['A2'].alignment = Alignment(horizontal='center')
    for col,h in enumerate(['Tanggal','NIK','Nama','Departemen','Jabatan','Jam Masuk','Jam Keluar','Jarak (m)','Status','Shift','Keterangan'],1):
        cell = ws.cell(4,col,h); cell.font = Font(bold=True,color='FFFFFF')
        cell.fill = PatternFill(fill_type='solid',fgColor='1E3A5F')
        cell.alignment = Alignment(horizontal='center')
    sc = {'hadir':'C8E6C9','telat':'FFE082','izin':'BBDEFB','alpha':'FFCDD2'}
    for ri,row in enumerate(data,5):
        for col,val in enumerate([str(row['tanggal']),row['nik'],row['nama'],
            row['departemen'] or '-',row['jabatan'] or '-',str(row['jam_masuk'] or '-'),
            str(row['jam_keluar'] or '-'),f"{row['jarak_masuk']:.0f}" if row['jarak_masuk'] else '-',
            row['status'],row['shift_nama'] or '-',row['keterangan'] or ''],1):
            cell = ws.cell(ri,col,val)
            cell.fill = PatternFill(fill_type='solid',fgColor=sc.get(row['status'],'FFFFFF'))
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"absensi_{bulan}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/export/pdf')
@admin_required
def export_pdf():
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT a.tanggal,u.nik,u.nama,u.departemen,a.jam_masuk,a.jam_keluar,
        a.jarak_masuk,a.status,s.nama as shift_nama
        FROM absensi a JOIN users u ON a.user_id=u.id LEFT JOIN shift s ON a.shift_id=s.id
        WHERE TO_CHAR(a.tanggal,'YYYY-MM')=%s ORDER BY a.tanggal,u.nama""", (bulan,))
    data = cur.fetchall()
    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    cur.close(); conn.close()
    out = io.BytesIO()
    doc = SimpleDocTemplate(out,pagesize=landscape(A4),rightMargin=1*cm,leftMargin=1*cm,topMargin=2*cm,bottomMargin=1*cm)
    el = [Paragraph(f"LAPORAN ABSENSI - {settings['nama_perusahaan'] if settings else 'PT Absensi'}",
              ParagraphStyle('T',fontSize=14,spaceAfter=4,fontName='Helvetica-Bold',alignment=1)),
          Paragraph(f"Periode: {bulan}", ParagraphStyle('S',fontSize=10,spaceAfter=10,alignment=1)),
          Spacer(1,0.3*cm)]
    td = [['No','Tanggal','NIK','Nama','Dept','Shift','Masuk','Keluar','Jarak','Status']]
    for i,row in enumerate(data,1):
        td.append([i,str(row['tanggal']),row['nik'],row['nama'],row['departemen'] or '-',
            row['shift_nama'] or '-',str(row['jam_masuk'] or '-'),str(row['jam_keluar'] or '-'),
            f"{row['jarak_masuk']:.0f}m" if row['jarak_masuk'] else '-',row['status']])
    t = Table(td,colWidths=[0.8*cm,2.2*cm,2.2*cm,3.5*cm,2.5*cm,2.5*cm,2*cm,2*cm,1.8*cm,1.8*cm])
    sc2 = {'hadir':colors.HexColor('#C8E6C9'),'telat':colors.HexColor('#FFE082'),'izin':colors.HexColor('#BBDEFB')}
    sty = [('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1E3A5F')),('TEXTCOLOR',(0,0),(-1,0),colors.white),
           ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),('FONTSIZE',(0,0),(-1,-1),7),
           ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F5F5F5')]),
           ('GRID',(0,0),(-1,-1),0.5,colors.grey),('ALIGN',(0,0),(-1,-1),'CENTER'),('VALIGN',(0,0),(-1,-1),'MIDDLE')]
    for i,row in enumerate(data,1):
        c = sc2.get(row['status'])
        if c: sty.append(('BACKGROUND',(9,i),(9,i),c))
    t.setStyle(TableStyle(sty)); el.append(t); doc.build(el); out.seek(0)
    return send_file(out,as_attachment=True,download_name=f"absensi_{bulan}.pdf",mimetype='application/pdf')

# ── LAPORAN PER PEGAWAI ───────────────────────────────────────────────────────
@app.route('/admin/laporan/pegawai')
@admin_required
def laporan_per_pegawai():
    conn = get_db(); cur = q(conn)
    bulan     = request.args.get('bulan', date.today().strftime('%Y-%m'))
    dept_id   = request.args.get('dept_id', '')
    status_f  = request.args.get('status_filter', '')

    # Daftar departemen untuk filter
    cur.execute("SELECT id, nama FROM departemen WHERE aktif=1 ORDER BY nama")
    daftar_dept = cur.fetchall()

    # Query pegawai
    where = ["u.role='user'", "u.status='active'"]
    params = []
    if dept_id:
        where.append("u.departemen_id=%s"); params.append(dept_id)
    cur.execute(f"""SELECT u.id,u.nik,u.nama,u.jabatan,u.departemen,u.foto,
        d.nama as dept_nama, s.nama as shift_nama, s.jam_masuk as shift_masuk, s.jam_keluar as shift_keluar
        FROM users u
        LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id
        WHERE {' AND '.join(where)} ORDER BY u.nama""", params)
    users = cur.fetchall()

    # Hitung hari kerja dalam bulan (Senin-Sabtu kecuali Minggu)
    y, m   = map(int, bulan.split('-'))
    import calendar
    total_hari_kerja = sum(
        1 for d2 in range(1, calendar.monthrange(y, m)[1]+1)
        if date(y, m, d2).weekday() < 6  # 0=Senin..5=Sabtu, 6=Minggu
    )

    rekap = []
    for u in users:
        cur.execute("""
            SELECT tanggal, jam_masuk, jam_keluar, status, keterangan,
                   jarak_masuk, shift_id
            FROM absensi
            WHERE user_id=%s AND TO_CHAR(tanggal,'YYYY-MM')=%s
            ORDER BY tanggal
        """, (u['id'], bulan))
        absensi_list = cur.fetchall()

        hadir  = sum(1 for a in absensi_list if a['status'] == 'hadir')
        telat  = sum(1 for a in absensi_list if a['status'] == 'telat')
        izin   = sum(1 for a in absensi_list if a['status'] == 'izin')
        alpha  = sum(1 for a in absensi_list if a['status'] == 'alpha')
        total_masuk = hadir + telat
        pct    = round(total_masuk / total_hari_kerja * 100) if total_hari_kerja else 0
        alpha_real = max(0, total_hari_kerja - hadir - telat - izin - alpha)

        row = {
            'user'        : u,
            'hadir'       : hadir,
            'telat'       : telat,
            'izin'        : izin,
            'alpha'       : alpha + alpha_real,
            'total_masuk' : total_masuk,
            'pct'         : pct,
            'hari_kerja'  : total_hari_kerja,
            'absensi'     : absensi_list,
        }
        if status_f == 'baik'   and pct <  90: continue
        if status_f == 'cukup'  and not (75 <= pct < 90): continue
        if status_f == 'kurang' and pct >= 75: continue
        rekap.append(row)

    cur.close(); conn.close()
    return render_template('admin/laporan_pegawai.html',
        rekap=rekap, bulan=bulan, daftar_dept=daftar_dept,
        dept_id=dept_id, status_filter=status_f,
        total_hari_kerja=total_hari_kerja)


@app.route('/admin/laporan/pegawai/<int:uid>')
@admin_required
def laporan_detail_pegawai(uid):
    conn = get_db(); cur = q(conn)
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))

    cur.execute("""SELECT u.*,d.nama as dept_nama,s.nama as shift_nama,
        s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar,s.toleransi_menit
        FROM users u
        LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id
        WHERE u.id=%s""", (uid,))
    user = cur.fetchone()
    if not user:
        cur.close(); conn.close()
        flash('Pegawai tidak ditemukan', 'error')
        return redirect(url_for('laporan_per_pegawai'))

    cur.execute("""
        SELECT a.*, s.nama as shift_nama, s.jam_masuk as sft_masuk, s.jam_keluar as sft_keluar
        FROM absensi a
        LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.user_id=%s AND TO_CHAR(a.tanggal,'YYYY-MM')=%s
        ORDER BY a.tanggal
    """, (uid, bulan))
    absensi_list = cur.fetchall()

    # Izin bulan ini
    cur.execute("""SELECT * FROM izin WHERE user_id=%s
        AND TO_CHAR(tanggal_mulai,'YYYY-MM')=%s ORDER BY tanggal_mulai""",
        (uid, bulan))
    izin_list = cur.fetchall()

    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    cur.close(); conn.close()

    # Statistik
    hadir  = sum(1 for a in absensi_list if a['status'] == 'hadir')
    telat  = sum(1 for a in absensi_list if a['status'] == 'telat')
    izin_c = sum(1 for a in absensi_list if a['status'] == 'izin')
    alpha  = sum(1 for a in absensi_list if a['status'] == 'alpha')

    import calendar
    y, m = map(int, bulan.split('-'))
    total_hari_kerja = sum(
        1 for d2 in range(1, calendar.monthrange(y, m)[1]+1)
        if date(y, m, d2).weekday() < 6
    )
    total_masuk = hadir + telat
    pct = round(total_masuk / total_hari_kerja * 100) if total_hari_kerja else 0

    # Build kalender
    cal_data = {}
    for a in absensi_list:
        cal_data[str(a['tanggal'])] = a

    # Buat list hari dalam bulan
    bulan_days = []
    for d2 in range(1, calendar.monthrange(y, m)[1]+1):
        tgl = date(y, m, d2)
        tgl_str = tgl.isoformat()
        bulan_days.append({
            'tanggal' : tgl,
            'hari'    : tgl.strftime('%a'),
            'is_minggu': tgl.weekday() == 6,
            'absensi' : cal_data.get(tgl_str),
        })

    return render_template('admin/laporan_detail_pegawai.html',
        user=user, bulan=bulan, absensi_list=absensi_list,
        izin_list=izin_list, hadir=hadir, telat=telat,
        izin_c=izin_c, alpha=alpha,
        total_masuk=total_masuk, total_hari_kerja=total_hari_kerja,
        pct=pct, bulan_days=bulan_days, settings=settings)


@app.route('/admin/laporan/pegawai/<int:uid>/export-excel')
@admin_required
def laporan_pegawai_export_excel(uid):
    conn = get_db(); cur = q(conn)
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))

    cur.execute("""SELECT u.*,d.nama as dept_nama,s.nama as shift_nama,
        s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=%s""", (uid,))
    user = cur.fetchone()

    cur.execute("""SELECT a.*,s.nama as shift_nama,s.jam_masuk as sft_masuk
        FROM absensi a LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.user_id=%s AND TO_CHAR(a.tanggal,'YYYY-MM')=%s
        ORDER BY a.tanggal""", (uid, bulan))
    rows = cur.fetchall()

    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    cur.close(); conn.close()

    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = f"Absensi {user['nama']}"

    # Column widths
    for col, w in zip('ABCDEFGHI', [12,10,12,12,14,14,14,12,25]):
        ws.column_dimensions[col].width = w

    navy  = 'FF1E3A5F'
    green = 'FFC8E6C9'; yellow = 'FFFFE082'; blue = 'FFBBDEFB'; red = 'FFFFCDD2'; grey = 'FFF5F5F5'

    def cell_style(cell, bold=False, bg=None, align='left', color='FF000000', size=10):
        cell.font = Font(bold=bold, size=size, color=color)
        if bg: cell.fill = PatternFill(fill_type='solid', fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

    instansi = settings['nama_perusahaan'] if settings else 'RST Slamet Riyadi'

    ws.merge_cells('A1:I1'); ws['A1'] = instansi.upper()
    cell_style(ws['A1'], bold=True, bg=navy, align='center', color='FFFFFFFF', size=12)
    ws.row_dimensions[1].height = 22

    ws.merge_cells('A2:I2'); ws['A2'] = f'LAPORAN ABSENSI PEGAWAI - {bulan}'
    cell_style(ws['A2'], bold=True, align='center', bg='FFE8EFF8', size=11)

    ws.merge_cells('A3:I3'); ws['A3'] = ''
    ws.row_dimensions[3].height = 6

    # Info pegawai
    info = [
        ('Nama', user['nama']),  ('NIK/NRP', user['nik'] or '-'),
        ('Jabatan', user['jabatan'] or '-'), ('Departemen', user['dept_nama'] or '-'),
        ('Shift', f"{user['shift_nama'] or '-'} ({user['shift_masuk'] or '-'} - {user['shift_keluar'] or '-'})"),
        ('Periode', bulan),
    ]
    for i, (lbl, val) in enumerate(info, 4):
        ws[f'A{i}'] = lbl; ws[f'B{i}'] = ':'
        ws.merge_cells(f'C{i}:I{i}'); ws[f'C{i}'] = val
        cell_style(ws[f'A{i}'], bold=True)
        cell_style(ws[f'C{i}'])
    ws.row_dimensions[9].height = 6

    # Header tabel
    headers = ['Tanggal', 'Hari', 'Jam Masuk', 'Jam Keluar', 'Shift', 'Jarak (m)', 'Status', 'Keterangan']
    for col, h in enumerate(headers, 1):
        c = ws.cell(10, col, h)
        cell_style(c, bold=True, bg=navy, align='center', color='FFFFFFFF')
    ws.row_dimensions[10].height = 18

    # Data
    sc = {'hadir': green, 'telat': yellow, 'izin': blue, 'alpha': red}
    for ri, row in enumerate(rows, 11):
        tgl = row['tanggal']
        hari = tgl.strftime('%A') if hasattr(tgl, 'strftime') else '-'
        bg_row = sc.get(row['status'], 'FFFFFFFF')
        vals = [
            str(tgl), hari,
            str(row['jam_masuk'] or '-'), str(row['jam_keluar'] or '-'),
            row['shift_nama'] or '-',
            f"{row['jarak_masuk']:.0f}" if row['jarak_masuk'] else '-',
            (row['status'] or '-').upper(), row['keterangan'] or '-'
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(ri, col, val)
            cell_style(c, bg=bg_row if col in (7,) else grey if ri % 2 == 0 else None,
                       align='center' if col in (1,2,3,4,6,7) else 'left')

    # Statistik
    hadir  = sum(1 for r in rows if r['status']=='hadir')
    telat  = sum(1 for r in rows if r['status']=='telat')
    izin   = sum(1 for r in rows if r['status']=='izin')
    alpha  = sum(1 for r in rows if r['status']=='alpha')

    sr = len(rows) + 12
    ws.row_dimensions[sr].height = 6
    stats = [
        ('Hadir', hadir, green), ('Telat', telat, yellow),
        ('Izin', izin, blue),  ('Alpha', alpha, red),
        ('Total Hadir+Telat', hadir+telat, 'FFFFFFFF'),
    ]
    for i, (lbl, val, bg) in enumerate(stats):
        r = sr + 1 + i
        ws[f'A{r}'] = lbl; ws[f'B{r}'] = val
        cell_style(ws[f'A{r}'], bold=True, bg=bg)
        cell_style(ws[f'B{r}'], bold=True, align='center', bg=bg)

    out = io.BytesIO(); wb.save(out); out.seek(0)
    safe_name = re.sub(r'[^\w]', '_', user['nama'])
    log_audit(conn if False else get_db(), 'EXPORT', 'laporan',
              deskripsi=f'Export Excel laporan pegawai {user["nama"]} bulan {bulan}')
    return send_file(out, as_attachment=True,
        download_name=f"absensi_{safe_name}_{bulan}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/admin/laporan/pegawai/<int:uid>/export-pdf')
@admin_required
def laporan_pegawai_export_pdf(uid):
    conn = get_db(); cur = q(conn)
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))

    cur.execute("""SELECT u.*,d.nama as dept_nama,s.nama as shift_nama,
        s.jam_masuk as shift_masuk,s.jam_keluar as shift_keluar
        FROM users u LEFT JOIN departemen d ON u.departemen_id=d.id
        LEFT JOIN shift s ON u.shift_id=s.id WHERE u.id=%s""", (uid,))
    user = cur.fetchone()

    cur.execute("""SELECT a.*,s.nama as shift_nama
        FROM absensi a LEFT JOIN shift s ON a.shift_id=s.id
        WHERE a.user_id=%s AND TO_CHAR(a.tanggal,'YYYY-MM')=%s
        ORDER BY a.tanggal""", (uid, bulan))
    rows = cur.fetchall()

    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    cur.close(); conn.close()

    hadir = sum(1 for r in rows if r['status']=='hadir')
    telat = sum(1 for r in rows if r['status']=='telat')
    izin  = sum(1 for r in rows if r['status']=='izin')
    alpha = sum(1 for r in rows if r['status']=='alpha')

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=2*cm, bottomMargin=1.5*cm)

    navy   = colors.HexColor('#1E3A5F')
    c_green= colors.HexColor('#C8E6C9')
    c_yell = colors.HexColor('#FFE082')
    c_blue = colors.HexColor('#BBDEFB')
    c_red  = colors.HexColor('#FFCDD2')
    c_grey = colors.HexColor('#F5F5F5')

    instansi = settings['nama_perusahaan'] if settings else 'RST Slamet Riyadi'

    el = []
    el.append(Paragraph(instansi.upper(),
        ParagraphStyle('H1', fontName='Helvetica-Bold', fontSize=13, alignment=1,
                       textColor=navy, spaceAfter=2)))
    el.append(Paragraph(f'LAPORAN ABSENSI PEGAWAI',
        ParagraphStyle('H2', fontName='Helvetica-Bold', fontSize=11, alignment=1, spaceAfter=2)))
    el.append(Paragraph(f'Periode: {bulan}',
        ParagraphStyle('H3', fontName='Helvetica', fontSize=9, alignment=1, spaceAfter=8,
                       textColor=colors.grey)))

    # Info pegawai
    info_data = [
        ['Nama', ':', user['nama'], 'NIK/NRP', ':', user['nik'] or '-'],
        ['Jabatan', ':', user['jabatan'] or '-', 'Departemen', ':', user['dept_nama'] or '-'],
        ['Shift', ':', f"{user['shift_nama'] or '-'} ({user['shift_masuk'] or '-'} - {user['shift_keluar'] or '-'})", '', '', ''],
    ]
    info_tbl = Table(info_data, colWidths=[2.2*cm,0.4*cm,5*cm,2.2*cm,0.4*cm,5*cm])
    info_tbl.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),'Helvetica'),
        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),
        ('FONTNAME',(3,0),(3,-1),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('BACKGROUND',(0,0),(-1,-1),colors.HexColor('#F0F4F8')),
        ('BOX',(0,0),(-1,-1),0.5,colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS',(0,0),(-1,-1),[colors.HexColor('#F0F4F8'), colors.HexColor('#E8EFF8')]),
    ]))
    el.append(info_tbl); el.append(Spacer(1, 0.4*cm))

    # Tabel absensi
    td = [['No','Tanggal','Hari','Jam Masuk','Jam Keluar','Shift','Jarak','Status','Keterangan']]
    for i, row in enumerate(rows, 1):
        tgl = row['tanggal']
        hari = tgl.strftime('%a') if hasattr(tgl, 'strftime') else '-'
        td.append([
            str(i), str(tgl), hari,
            str(row['jam_masuk'] or '-'), str(row['jam_keluar'] or '-'),
            row['shift_nama'] or '-',
            f"{row['jarak_masuk']:.0f}m" if row['jarak_masuk'] else '-',
            (row['status'] or '-').upper(),
            row['keterangan'] or '-'
        ])

    col_w = [0.7*cm,2.3*cm,1.3*cm,2.1*cm,2.1*cm,2.5*cm,1.4*cm,1.6*cm,3*cm]
    t = Table(td, colWidths=col_w, repeatRows=1)
    sty = [
        ('BACKGROUND',(0,0),(-1,0), navy),
        ('TEXTCOLOR',(0,0),(-1,0), colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),7.5),
        ('ALIGN',(0,0),(-1,0),'CENTER'),
        ('ALIGN',(0,1),(7,-1),'CENTER'),
        ('ALIGN',(8,1),(8,-1),'LEFT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('GRID',(0,0),(-1,-1),0.4, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, c_grey]),
        ('BOTTOMPADDING',(0,0),(-1,-1),4),
        ('TOPPADDING',(0,0),(-1,-1),4),
    ]
    sc_map = {'hadir': c_green, 'telat': c_yell, 'izin': c_blue, 'alpha': c_red}
    for i, row in enumerate(rows, 1):
        c = sc_map.get(row['status'])
        if c: sty.append(('BACKGROUND',(7,i),(7,i), c))
    t.setStyle(TableStyle(sty))
    el.append(t); el.append(Spacer(1, 0.4*cm))

    # Rekap statistik
    stat_data = [
        ['Hadir', str(hadir), 'Telat', str(telat), 'Izin', str(izin), 'Alpha', str(alpha)],
    ]
    st = Table(stat_data, colWidths=[2*cm,1.2*cm,2*cm,1.2*cm,2*cm,1.2*cm,2*cm,1.2*cm])
    st.setStyle(TableStyle([
        ('FONTNAME',(0,0),(-1,-1),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),9),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('BACKGROUND',(0,0),(1,0), c_green),
        ('BACKGROUND',(2,0),(3,0), c_yell),
        ('BACKGROUND',(4,0),(5,0), c_blue),
        ('BACKGROUND',(6,0),(7,0), c_red),
        ('BOX',(0,0),(-1,-1),0.5, colors.HexColor('#CBD5E1')),
        ('INNERGRID',(0,0),(-1,-1),0.3, colors.HexColor('#CBD5E1')),
        ('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('TOPPADDING',(0,0),(-1,-1),5),
    ]))
    el.append(st)
    doc.build(el); out.seek(0)

    safe_name = re.sub(r'[^\w]', '_', user['nama'])
    return send_file(out, as_attachment=True,
        download_name=f"absensi_{safe_name}_{bulan}.pdf",
        mimetype='application/pdf')


# ── ADMIN GRAFIK ──────────────────────────────────────────────────────────────
@app.route('/admin/grafik')
@admin_required
def admin_grafik():
    conn = get_db(); cur = q(conn)
    bulan = request.args.get('bulan', date.today().strftime('%Y-%m'))
    cur.execute("""SELECT tanggal::text,
        SUM(CASE WHEN status='hadir' THEN 1 ELSE 0 END) as hadir,
        SUM(CASE WHEN status='telat' THEN 1 ELSE 0 END) as telat,
        SUM(CASE WHEN status='izin' THEN 1 ELSE 0 END) as izin
        FROM absensi WHERE TO_CHAR(tanggal,'YYYY-MM')=%s
        GROUP BY tanggal ORDER BY tanggal""", (bulan,))
    harian = cur.fetchall()
    cur.execute("""SELECT d.nama as departemen,d.warna,
        SUM(CASE WHEN a.status='hadir' THEN 1 ELSE 0 END) as hadir,
        SUM(CASE WHEN a.status='telat' THEN 1 ELSE 0 END) as telat,
        SUM(CASE WHEN a.status='izin' THEN 1 ELSE 0 END) as izin,
        COUNT(a.id) as total
        FROM departemen d
        LEFT JOIN users u ON u.departemen_id=d.id AND u.role='user' AND u.status='active'
        LEFT JOIN absensi a ON u.id=a.user_id AND TO_CHAR(a.tanggal,'YYYY-MM')=%s
        WHERE d.aktif=1 GROUP BY d.id,d.nama,d.warna""", (bulan,))
    dept = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/grafik.html',
        harian=json.dumps([dict(r) for r in harian]),
        dept=json.dumps([dict(r) for r in dept]), bulan=bulan)

# ── ADMIN SETTINGS ────────────────────────────────────────────────────────────
@app.route('/admin/settings', methods=['GET','POST'])
@admin_required
def admin_settings():
    conn = get_db(); cur = q(conn)
    if request.method == 'POST':
        # Handle logo upload
        logo_filename = None
        logo_file = request.files.get('logo')
        if logo_file and logo_file.filename:
            ext = logo_file.filename.rsplit('.', 1)[-1].lower() if '.' in logo_file.filename else ''
            if ext in {'png', 'jpg', 'jpeg', 'svg', 'gif', 'webp'}:
                logo_dir = os.path.join('static', 'uploads', 'logo')
                os.makedirs(logo_dir, exist_ok=True)
                logo_filename = secure_filename(f"logo.{ext}")
                logo_file.save(os.path.join(logo_dir, logo_filename))

        if logo_filename:
            cur.execute("""UPDATE settings SET nama_perusahaan=%s,
                office_lat=%s,office_lng=%s,max_distance=%s,logo=%s WHERE id=1""",
                (request.form['nama_perusahaan'],
                 float(request.form['office_lat']), float(request.form['office_lng']),
                 int(request.form['max_distance']), logo_filename))
        else:
            cur.execute("""UPDATE settings SET nama_perusahaan=%s,
                office_lat=%s,office_lng=%s,max_distance=%s WHERE id=1""",
                (request.form['nama_perusahaan'],
                 float(request.form['office_lat']), float(request.form['office_lng']),
                 int(request.form['max_distance'])))
        conn.commit()
        log_audit(conn, 'SETTING', 'settings', deskripsi='Update pengaturan sistem',
            data_baru={'nama_perusahaan':request.form.get('nama_perusahaan'),
                       'office_lat':request.form.get('office_lat'),
                       'office_lng':request.form.get('office_lng'),
                       'max_distance':request.form.get('max_distance')})
        flash('Settings disimpan!', 'success')
    cur.execute("SELECT * FROM settings WHERE id=1")
    settings = cur.fetchone()
    # Ambil data pengumuman untuk tab pengumuman
    pengumuman = None
    try:
        cur.execute("SELECT * FROM pengumuman ORDER BY id LIMIT 1")
        row = cur.fetchone()
        pengumuman = dict(row) if row else None
    except Exception:
        pass
    cur.close(); conn.close()
    return render_template('admin/settings.html', settings=settings, pengumuman=pengumuman)

@app.route('/admin/settings/ganti-password', methods=['POST'])
@admin_required
def admin_ganti_password_sendiri():
    uid = session['user_id']
    pw_lama     = request.form.get('password_lama', '')
    pw_baru     = request.form.get('password_baru', '')
    pw_konfirm  = request.form.get('password_konfirm', '')
    if not pw_baru or len(pw_baru) < 6:
        flash('Password baru minimal 6 karakter.', 'error')
        return redirect(url_for('admin_settings') + '#keamanan')
    if pw_baru != pw_konfirm:
        flash('Konfirmasi password tidak cocok.', 'error')
        return redirect(url_for('admin_settings') + '#keamanan')
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT password FROM users WHERE id=%s", (uid,))
    user = cur.fetchone()
    if not user or not check_password_hash(user['password'], pw_lama):
        flash('Password lama salah.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('admin_settings') + '#keamanan')
    cur.execute("UPDATE users SET password=%s WHERE id=%s", (generate_password_hash(pw_baru), uid))
    conn.commit()
    log_audit(conn, 'PASSWORD', 'settings',
        deskripsi=f'Admin ganti password: {session.get("nama")}',
        ref_id=uid, ref_table='users')
    cur.close(); conn.close()
    flash('Password admin berhasil diubah!', 'success')
    return redirect(url_for('admin_settings') + '#keamanan')

@app.route('/admin/settings/hapus-logo', methods=['POST'])
@admin_required
def hapus_logo():
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT logo FROM settings WHERE id=1")
    s = cur.fetchone()
    if s and s['logo']:
        path = os.path.join('static', 'uploads', 'logo', s['logo'])
        if os.path.exists(path):
            os.remove(path)
        cur.execute("UPDATE settings SET logo=NULL WHERE id=1")
        conn.commit()
        flash('Logo berhasil dihapus.', 'success')
    cur.close(); conn.close()
    return redirect(url_for('admin_settings'))

# ── API ───────────────────────────────────────────────────────────────────────
@app.route('/api/shift-by-dept/<int:dept_id>')
@login_required
def api_shift_by_dept(dept_id):
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT s.* FROM shift s JOIN departemen_shift ds ON s.id=ds.shift_id
        WHERE ds.departemen_id=%s AND s.aktif=1""", (dept_id,))
    shifts = cur.fetchall()
    cur.close(); conn.close()
    return jsonify([dict(s) for s in shifts])

# ── UBAH PASSWORD ─────────────────────────────────────────────────────────────

@app.route('/ubah-password', methods=['POST'])
@login_required
def ubah_password():
    uid = session['user_id']
    pw_lama = request.form.get('password_lama','')
    pw_baru = request.form.get('password_baru','')
    pw_konfirm = request.form.get('password_konfirm','')
    if not pw_baru or len(pw_baru) < 6:
        flash('Password baru minimal 6 karakter.', 'error')
        return redirect(url_for('profil'))
    if pw_baru != pw_konfirm:
        flash('Konfirmasi password tidak cocok.', 'error')
        return redirect(url_for('profil'))
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT password FROM users WHERE id=%s", (uid,))
    user = cur.fetchone()
    if not user or not check_password_hash(user['password'], pw_lama):
        flash('Password lama salah.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('profil'))
    cur.execute("UPDATE users SET password=%s WHERE id=%s", (generate_password_hash(pw_baru), uid))
    conn.commit()
    log_audit(conn, 'PASSWORD', 'profil',
        deskripsi=f'Ubah password: {session.get("nama")}',
        ref_id=uid, ref_table='users')
    cur.close(); conn.close()
    flash('Password berhasil diubah!', 'success')
    return redirect(url_for('profil'))

@app.route('/admin/pegawai/<int:uid>/ubah-password', methods=['POST'])
@admin_required
def admin_ubah_password(uid):
    pw_baru = request.form.get('password_baru','')
    pw_konfirm = request.form.get('password_konfirm','')
    if not pw_baru or len(pw_baru) < 6:
        flash('Password baru minimal 6 karakter.', 'error')
        return redirect(url_for('edit_pegawai', uid=uid))
    if pw_baru != pw_konfirm:
        flash('Konfirmasi password tidak cocok.', 'error')
        return redirect(url_for('edit_pegawai', uid=uid))
    conn = get_db(); cur = q(conn)
    cur.execute("UPDATE users SET password=%s WHERE id=%s", (generate_password_hash(pw_baru), uid))
    conn.commit(); cur.close(); conn.close()
    flash('Password pegawai berhasil diubah!', 'success')
    return redirect(url_for('edit_pegawai', uid=uid))

# ── PEJABAT TTD ───────────────────────────────────────────────────────────────

@app.route('/admin/pejabat-ttd')
@admin_required
def admin_pejabat_ttd():
    conn = get_db(); cur = q(conn)
    # Buat tabel jika belum ada
    cur2 = conn.cursor()
    cur2.execute("""CREATE TABLE IF NOT EXISTS pejabat_ttd (
        id SERIAL PRIMARY KEY, nama TEXT NOT NULL, jabatan TEXT NOT NULL,
        nip TEXT, departemen_id INTEGER REFERENCES departemen(id) ON DELETE SET NULL,
        level_approval INTEGER NOT NULL DEFAULT 1, role_label TEXT NOT NULL DEFAULT 'Pejabat',
        ttd_file TEXT, aktif INTEGER DEFAULT 1, urutan INTEGER DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit(); cur2.close()
    cur.execute("""SELECT p.*, d.nama as nama_dept FROM pejabat_ttd p
        LEFT JOIN departemen d ON p.departemen_id=d.id ORDER BY p.level_approval, p.urutan""")
    pejabat_list = cur.fetchall()
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/pejabat_ttd.html', pejabat_list=pejabat_list, depts=depts)

@app.route('/admin/pejabat-ttd/tambah', methods=['POST'])
@admin_required
def admin_pejabat_ttd_tambah():
    nama=request.form.get('nama','').strip(); jabatan=request.form.get('jabatan','').strip()
    nip=request.form.get('nip','').strip(); dept_id=request.form.get('departemen_id') or None
    level=int(request.form.get('level_approval',1)); role_label=request.form.get('role_label','').strip() or jabatan
    urutan=int(request.form.get('urutan',0))
    ttd_file=None; f=request.files.get('ttd_file')
    if f and f.filename:
        ext=f.filename.rsplit('.',1)[-1].lower()
        if ext in {'png','jpg','jpeg'}:
            os.makedirs(app.config['TTD_FOLDER'],exist_ok=True)
            fn=secure_filename(f"pjb_{nama.replace(' ','_')}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
            f.save(os.path.join(app.config['TTD_FOLDER'],fn)); ttd_file=fn
    conn=get_db(); cur=q(conn)
    cur.execute("""INSERT INTO pejabat_ttd (nama,jabatan,nip,departemen_id,level_approval,role_label,ttd_file,urutan)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""", (nama,jabatan,nip,dept_id,level,role_label,ttd_file,urutan))
    conn.commit(); cur.close(); conn.close()
    flash(f'Pejabat {nama} berhasil ditambahkan.','success')
    return redirect(url_for('admin_pejabat_ttd'))

@app.route('/admin/pejabat-ttd/<int:pid>/edit', methods=['POST'])
@admin_required
def admin_pejabat_ttd_edit(pid):
    nama=request.form.get('nama','').strip(); jabatan=request.form.get('jabatan','').strip()
    nip=request.form.get('nip','').strip(); dept_id=request.form.get('departemen_id') or None
    level=int(request.form.get('level_approval',1)); role_label=request.form.get('role_label','').strip() or jabatan
    urutan=int(request.form.get('urutan',0)); aktif=int(request.form.get('aktif',1))
    conn=get_db(); cur=q(conn)
    f=request.files.get('ttd_file')
    if f and f.filename:
        ext=f.filename.rsplit('.',1)[-1].lower()
        if ext in {'png','jpg','jpeg'}:
            os.makedirs(app.config['TTD_FOLDER'],exist_ok=True)
            fn=secure_filename(f"pjb_{pid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
            f.save(os.path.join(app.config['TTD_FOLDER'],fn))
            cur.execute("""UPDATE pejabat_ttd SET nama=%s,jabatan=%s,nip=%s,departemen_id=%s,
                level_approval=%s,role_label=%s,ttd_file=%s,urutan=%s,aktif=%s WHERE id=%s""",
                (nama,jabatan,nip,dept_id,level,role_label,fn,urutan,aktif,pid))
        else:
            cur.execute("""UPDATE pejabat_ttd SET nama=%s,jabatan=%s,nip=%s,departemen_id=%s,
                level_approval=%s,role_label=%s,urutan=%s,aktif=%s WHERE id=%s""",
                (nama,jabatan,nip,dept_id,level,role_label,urutan,aktif,pid))
    else:
        cur.execute("""UPDATE pejabat_ttd SET nama=%s,jabatan=%s,nip=%s,departemen_id=%s,
            level_approval=%s,role_label=%s,urutan=%s,aktif=%s WHERE id=%s""",
            (nama,jabatan,nip,dept_id,level,role_label,urutan,aktif,pid))
    conn.commit(); cur.close(); conn.close()
    flash('Data pejabat berhasil diperbarui.','success')
    return redirect(url_for('admin_pejabat_ttd'))

@app.route('/admin/pejabat-ttd/<int:pid>/hapus', methods=['POST'])
@admin_required
def admin_pejabat_ttd_hapus(pid):
    conn=get_db(); cur=q(conn)
    cur.execute("DELETE FROM pejabat_ttd WHERE id=%s",(pid,))
    conn.commit(); cur.close(); conn.close()
    flash('Pejabat berhasil dihapus.','success')
    return redirect(url_for('admin_pejabat_ttd'))

@app.route('/api/pejabat-ttd')
@login_required
def api_pejabat_ttd():
    level=request.args.get('level',type=int)
    conn=get_db(); cur=q(conn)
    if level:
        cur.execute("""SELECT id,nama,jabatan,role_label,ttd_file FROM pejabat_ttd
            WHERE aktif=1 AND level_approval=%s ORDER BY urutan""",(level,))
    else:
        cur.execute("""SELECT id,nama,jabatan,role_label,level_approval,ttd_file
            FROM pejabat_ttd WHERE aktif=1 ORDER BY level_approval,urutan""")
    data=cur.fetchall(); cur.close(); conn.close()
    return jsonify(data)



# ══════════════════════════════════════════════════════════════════════════════
# ── E-DOSIR USER ──────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/dosir')
@login_required
def dosir():
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM users WHERE id=%s", (uid,))
    user = cur.fetchone()
    dept_id = user['departemen_id'] if user else None

    # Jenis dokumen yang perlu diupload (global + dept)
    if dept_id:
        cur.execute("""SELECT * FROM dosir_jenis
            WHERE aktif=1 AND (departemen_id IS NULL OR departemen_id=%s)
            ORDER BY urutan, nama""", (dept_id,))
    else:
        cur.execute("SELECT * FROM dosir_jenis WHERE aktif=1 AND departemen_id IS NULL ORDER BY urutan, nama")
    jenis_list = cur.fetchall()

    # File yang sudah diupload user
    cur.execute("SELECT * FROM dosir_file WHERE user_id=%s", (uid,))
    uploads = {r['jenis_id']: dict(r) for r in cur.fetchall()}

    cur.close(); conn.close()
    return render_template('dosir.html', jenis_list=jenis_list, uploads=uploads, user=user, today=date.today())


@app.route('/dosir/upload/<int:jid>', methods=['POST'])
@login_required
def dosir_upload(jid):
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM dosir_jenis WHERE id=%s AND aktif=1", (jid,))
    jenis = cur.fetchone()
    if not jenis:
        flash('Jenis dokumen tidak ditemukan.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('dosir'))

    f = request.files.get('file')
    keterangan = request.form.get('keterangan', '').strip()
    tgl_expired_raw = request.form.get('tanggal_expired', '').strip()
    tanggal_expired = tgl_expired_raw if tgl_expired_raw else None
    if not f or not f.filename:
        flash('Pilih file terlebih dahulu.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('dosir'))

    ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
    if ext not in ALLOWED_DOSIR:
        flash('Format file tidak diizinkan. Gunakan PDF, JPG, atau PNG.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('dosir'))

    dosir_folder = app.config['DOSIR_FOLDER']
    os.makedirs(dosir_folder, exist_ok=True)
    original = f.filename
    fn = secure_filename(f"dosir_{uid}_{jid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
    f.save(os.path.join(dosir_folder, fn))

    cur.execute("""INSERT INTO dosir_file (user_id, jenis_id, filename, original_name, keterangan, tanggal_expired, status)
        VALUES (%s,%s,%s,%s,%s,%s,'pending')
        ON CONFLICT (user_id, jenis_id) DO UPDATE
        SET filename=EXCLUDED.filename, original_name=EXCLUDED.original_name,
            keterangan=EXCLUDED.keterangan, tanggal_expired=EXCLUDED.tanggal_expired,
            status='pending', catatan_admin=NULL, uploaded_at=CURRENT_TIMESTAMP, verified_at=NULL""",
        (uid, jid, fn, original, keterangan, tanggal_expired))
    conn.commit()
    cur.close(); conn.close()
    flash(f'Dokumen "{jenis["nama"]}" berhasil diupload. Menunggu verifikasi admin.', 'success')
    return redirect(url_for('dosir'))


@app.route('/dosir/file/<int:fid>')
@login_required
def dosir_view(fid):
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM dosir_file WHERE id=%s AND user_id=%s", (fid, uid))
    df = cur.fetchone()
    cur.close(); conn.close()
    if not df:
        flash('File tidak ditemukan.', 'error')
        return redirect(url_for('dosir'))
    path = os.path.join(app.config['DOSIR_FOLDER'], df['filename'])
    return send_file(path, as_attachment=False)


# ══════════════════════════════════════════════════════════════════════════════
# ── E-DOSIR ADMIN ─────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/dosir')
@admin_required
def admin_dosir():
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    cur.execute("""SELECT dj.*, d.nama as dept_nama
        FROM dosir_jenis dj LEFT JOIN departemen d ON dj.departemen_id=d.id
        WHERE dj.aktif=1 ORDER BY dj.urutan, dj.nama""")
    jenis_list = cur.fetchall()
    # Statistik upload per jenis
    cur.execute("""SELECT jenis_id,
        COUNT(*) as total,
        SUM(CASE WHEN status='verified' THEN 1 ELSE 0 END) as verified,
        SUM(CASE WHEN status='pending' THEN 1 ELSE 0 END) as pending,
        SUM(CASE WHEN status='rejected' THEN 1 ELSE 0 END) as rejected,
        SUM(CASE WHEN tanggal_expired IS NOT NULL AND tanggal_expired < CURRENT_DATE THEN 1 ELSE 0 END) as expired
        FROM dosir_file GROUP BY jenis_id""")
    stats = {r['jenis_id']: dict(r) for r in cur.fetchall()}
    cur.close(); conn.close()
    return render_template('admin/dosir.html', depts=depts, jenis_list=jenis_list, stats=stats)


@app.route('/admin/dosir/jenis/tambah', methods=['POST'])
@admin_required
def admin_dosir_tambah():
    nama = request.form.get('nama','').strip()
    deskripsi = request.form.get('deskripsi','').strip()
    wajib = 1 if request.form.get('wajib') else 0
    dept_id = request.form.get('departemen_id') or None
    urutan = request.form.get('urutan', 0)
    if not nama:
        flash('Nama dokumen tidak boleh kosong.', 'error')
        return redirect(url_for('admin_dosir'))
    conn = get_db(); cur = q(conn)
    cur.execute("""INSERT INTO dosir_jenis (nama,deskripsi,wajib,departemen_id,urutan)
        VALUES (%s,%s,%s,%s,%s)""", (nama, deskripsi, wajib, dept_id, urutan))
    conn.commit(); cur.close(); conn.close()
    flash(f'Jenis dokumen "{nama}" berhasil ditambahkan.', 'success')
    return redirect(url_for('admin_dosir'))


@app.route('/admin/dosir/jenis/edit/<int:jid>', methods=['POST'])
@admin_required
def admin_dosir_edit(jid):
    nama = request.form.get('nama','').strip()
    deskripsi = request.form.get('deskripsi','').strip()
    wajib = 1 if request.form.get('wajib') else 0
    dept_id = request.form.get('departemen_id') or None
    urutan = request.form.get('urutan', 0)
    conn = get_db(); cur = q(conn)
    cur.execute("""UPDATE dosir_jenis SET nama=%s,deskripsi=%s,wajib=%s,departemen_id=%s,urutan=%s
        WHERE id=%s""", (nama, deskripsi, wajib, dept_id, urutan, jid))
    conn.commit(); cur.close(); conn.close()
    flash('Jenis dokumen berhasil diperbarui.', 'success')
    return redirect(url_for('admin_dosir'))


@app.route('/admin/dosir/jenis/hapus/<int:jid>', methods=['POST'])
@admin_required
def admin_dosir_hapus(jid):
    conn = get_db(); cur = q(conn)
    cur.execute("UPDATE dosir_jenis SET aktif=0 WHERE id=%s", (jid,))
    conn.commit(); cur.close(); conn.close()
    flash('Jenis dokumen berhasil dihapus.', 'success')
    return redirect(url_for('admin_dosir'))


@app.route('/admin/dosir/files')
@admin_required
def admin_dosir_files():
    conn = get_db(); cur = q(conn)
    dept_id = request.args.get('dept_id')
    status_filter = request.args.get('status', '')
    cur.execute("SELECT * FROM departemen WHERE aktif=1 ORDER BY nama")
    depts = cur.fetchall()
    query = """SELECT df.*,
        u.nama as user_nama, u.nik, d.nama as dept_nama,
        dj.nama as jenis_nama, dj.wajib,
        CASE WHEN df.tanggal_expired IS NOT NULL AND df.tanggal_expired < CURRENT_DATE THEN TRUE ELSE FALSE END as is_expired
        FROM dosir_file df
        JOIN users u ON df.user_id=u.id
        LEFT JOIN departemen d ON u.departemen_id=d.id
        JOIN dosir_jenis dj ON df.jenis_id=dj.id
        WHERE 1=1"""
    params = []
    if dept_id:
        query += " AND u.departemen_id=%s"; params.append(dept_id)
    if status_filter == 'expired':
        query += " AND df.tanggal_expired IS NOT NULL AND df.tanggal_expired < CURRENT_DATE"
    elif status_filter:
        query += " AND df.status=%s"; params.append(status_filter)
    query += " ORDER BY df.uploaded_at DESC"
    cur.execute(query, params)
    files = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/dosir_files.html', files=files, depts=depts,
        dept_id=dept_id, status_filter=status_filter, today=date.today())


@app.route('/admin/dosir/verify/<int:fid>/<action>', methods=['POST'])
@admin_required
def admin_dosir_verify(fid, action):
    catatan = request.form.get('catatan','').strip()
    tgl_expired_raw = request.form.get('tanggal_expired','').strip()
    tanggal_expired = tgl_expired_raw if tgl_expired_raw else None
    conn = get_db(); cur = q(conn)
    if action == 'verify':
        cur.execute("""UPDATE dosir_file SET status='verified', catatan_admin=%s,
            tanggal_expired=COALESCE(%s::DATE, tanggal_expired),
            verified_at=CURRENT_TIMESTAMP WHERE id=%s""", (catatan, tanggal_expired, fid))
        flash('Dokumen berhasil diverifikasi.', 'success')
    elif action == 'reject':
        cur.execute("""UPDATE dosir_file SET status='rejected', catatan_admin=%s,
            tanggal_expired=COALESCE(%s::DATE, tanggal_expired),
            verified_at=CURRENT_TIMESTAMP WHERE id=%s""", (catatan, tanggal_expired, fid))
        flash('Dokumen ditolak.', 'info')
    conn.commit(); cur.close(); conn.close()
    return redirect(url_for('admin_dosir_files', **request.args))


@app.route('/admin/dosir/file/<int:fid>')
@admin_required
def admin_dosir_view(fid):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM dosir_file WHERE id=%s", (fid,))
    df = cur.fetchone()
    cur.close(); conn.close()
    if not df:
        flash('File tidak ditemukan.', 'error')
        return redirect(url_for('admin_dosir_files'))
    path = os.path.join(app.config['DOSIR_FOLDER'], df['filename'])
    return send_file(path, as_attachment=False)


# ══════════════════════════════════════════════════════════════════════════════
# ── HELPER NOTIFIKASI ─────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

def kirim_notif(conn, user_id, judul, pesan, tipe='info', ref_id=None, ref_type=None):
    cur = conn.cursor()
    cur.execute("""INSERT INTO notifikasi (user_id,judul,pesan,tipe,ref_id,ref_type)
        VALUES (%s,%s,%s,%s,%s,%s)""", (user_id, judul, pesan, tipe, ref_id, ref_type))
    cur.close()

@app.route('/api/notifikasi')
@login_required
def api_notifikasi():
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT * FROM notifikasi WHERE user_id=%s ORDER BY created_at DESC LIMIT 20""", (uid,))
    notifs = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT COUNT(*) as c FROM notifikasi WHERE user_id=%s AND dibaca=0", (uid,))
    unread = cur.fetchone()['c']
    cur.close(); conn.close()
    return jsonify({'notifs': notifs, 'unread': unread})

@app.route('/notifikasi/baca/<int:nid>', methods=['POST'])
@login_required
def baca_notif(nid):
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("UPDATE notifikasi SET dibaca=1 WHERE id=%s AND user_id=%s", (nid, uid))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/notifikasi/baca-semua', methods=['POST'])
@login_required
def baca_semua_notif():
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("UPDATE notifikasi SET dibaca=1 WHERE user_id=%s", (uid,))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

@app.route('/notifikasi')
@login_required
def halaman_notifikasi():
    uid = session['user_id']
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT * FROM notifikasi WHERE user_id=%s ORDER BY created_at DESC LIMIT 50""", (uid,))
    notifs = [dict(r) for r in cur.fetchall()]
    # Tandai semua sebagai dibaca saat halaman dibuka
    cur.execute("UPDATE notifikasi SET dibaca=1 WHERE user_id=%s", (uid,))
    conn.commit(); cur.close(); conn.close()
    return render_template('notifikasi.html', notifs=notifs)


# ══════════════════════════════════════════════════════════════════════════════
# ── MODUL ARSIP ───────────────────────────────────────────────────────────────
# ══════════════════════════════════════════════════════════════════════════════

JENIS_ARSIP = [
    'Surat Keputusan', 'Surat Edaran', 'Surat Perintah', 'Surat Tugas',
    'Nota Dinas', 'Berita Acara', 'Surat Undangan', 'Surat Keterangan',
    'Surat Permohonan', 'Surat Perjanjian', 'Memo', 'SOP', 'Surat Telegram', 'Lainnya',
]

JENIS_BISA_BAGIKAN = ['Surat Perintah', 'Surat Tugas', 'Surat Telegram']

KATEGORI_ARSIP = [
    'Pokja TKRS', 'PMKP', 'KPS', 'MFK', 'PPI', 'PPK',
    'MRMIK', 'PROGNAS', 'AKP', 'PP', 'PAP', 'PAB',
    'PKPO', 'HPK', 'KE', 'SKP',
]


def _kirim_wa_arsip(cfg, no_hp, nama_user, judul_arsip, jenis):
    if not cfg.get('fonnte_token'):
        return False, "Token Fonnte belum dikonfigurasi."
    nomor = re.sub(r'\D', '', no_hp)
    if nomor.startswith('0'):        nomor = '62' + nomor[1:]
    elif not nomor.startswith('62'): nomor = '62' + nomor
    instansi = cfg.get('nama_perusahaan', 'Instansi')
    pesan = (
        f"\U0001f4cb *{jenis} \u2014 {instansi}*\n\n"
        f"Halo *{nama_user}*,\n\n"
        f"Anda mendapat distribusi surat:\n"
        f"*{judul_arsip}*\n\n"
        f"Silakan buka aplikasi SIAP untuk melihat dokumen.\n\n"
        f"_Pesan ini dikirim otomatis oleh sistem._"
    )
    try:
        payload = json.dumps({'target': nomor, 'message': pesan, 'countryCode': '62'}).encode()
        req = _urllib_req.Request(
            cfg['fonnte_url'], data=payload,
            headers={'Authorization': cfg['fonnte_token'], 'Content-Type': 'application/json'},
            method='POST')
        with _urllib_req.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode())
            if result.get('status') in (True, 'true'): return True, ''
            return False, str(result.get('reason') or result.get('message') or result)
    except Exception as e:
        return False, str(e)

@app.route('/arsip')
@login_required
def arsip_user():
    conn = get_db(); cur = q(conn)
    jenis_filter    = request.args.get('jenis', '')
    kategori_filter = request.args.get('kategori', '')
    q_str           = request.args.get('q', '')
    query = """SELECT a.*, u.nama as pembuat_nama
        FROM arsip a JOIN users u ON a.dibuat_oleh=u.id
        WHERE a.aktif=1"""
    params = []
    if jenis_filter:
        query += " AND a.jenis=%s"; params.append(jenis_filter)
    if kategori_filter:
        query += " AND a.kategori=%s"; params.append(kategori_filter)
    if q_str:
        query += " AND (a.judul ILIKE %s OR a.nomor ILIKE %s OR a.keterangan ILIKE %s)"
        params += [f'%{q_str}%'] * 3
    query += " ORDER BY a.tanggal DESC, a.created_at DESC"
    cur.execute(query, params)
    arsip_list = cur.fetchall()
    cur.close(); conn.close()
    return render_template('arsip.html', arsip_list=arsip_list,
        jenis_list=JENIS_ARSIP, kategori_list=KATEGORI_ARSIP,
        jenis_filter=jenis_filter, kategori_filter=kategori_filter, q=q_str)


@app.route('/arsip/upload', methods=['GET', 'POST'])
@login_required
def arsip_upload_user():
    uid = session['user_id']
    if request.method == 'POST':
        judul      = request.form.get('judul', '').strip()
        jenis      = request.form.get('jenis', '').strip()
        kategori   = request.form.get('kategori', '').strip()
        nomor      = request.form.get('nomor', '').strip()
        tanggal    = request.form.get('tanggal', date.today().isoformat())
        keterangan = request.form.get('keterangan', '').strip()
        if not judul or not jenis:
            flash('Judul dan jenis wajib diisi.', 'error')
            return redirect(url_for('arsip_upload_user'))
        f = request.files.get('file')
        if not f or not f.filename:
            flash('File PDF wajib diupload.', 'error')
            return redirect(url_for('arsip_upload_user'))
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ALLOWED_DOSIR:
            flash('Format tidak diizinkan. Gunakan PDF, JPG, atau PNG.', 'error')
            return redirect(url_for('arsip_upload_user'))
        os.makedirs(app.config['SURAT_FOLDER'], exist_ok=True)
        original_name = f.filename
        filename = secure_filename(f"arsip_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
        f.save(os.path.join(app.config['SURAT_FOLDER'], filename))
        conn = get_db(); cur = q(conn)
        cur.execute("""INSERT INTO arsip
            (judul, jenis, kategori, nomor, tanggal, keterangan, filename, original_name, dibuat_oleh, sumber)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,'user') RETURNING id""",
            (judul, jenis, kategori or None, nomor or None, tanggal,
             keterangan or None, filename, original_name, uid))
        aid = cur.fetchone()['id']
        log_audit(conn, 'UPLOAD', 'arsip', deskripsi=f'User upload arsip: {judul}', ref_id=aid, ref_table='arsip')
        conn.commit(); cur.close(); conn.close()
        flash(f'Arsip "{judul}" berhasil diupload.', 'success')
        return redirect(url_for('arsip_user'))
    return render_template('arsip_upload.html',
        jenis_list=JENIS_ARSIP, kategori_list=KATEGORI_ARSIP, today=date.today())


@app.route('/arsip/<int:aid>')
@login_required
def arsip_detail(aid):
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT a.*, u.nama as pembuat_nama, u.jabatan as pembuat_jabatan
        FROM arsip a JOIN users u ON a.dibuat_oleh=u.id
        WHERE a.id=%s AND a.aktif=1""", (aid,))
    arsip = cur.fetchone()
    cur.close(); conn.close()
    if not arsip:
        flash('Arsip tidak ditemukan.', 'error')
        return redirect(url_for('arsip_user'))
    return render_template('arsip_detail.html', arsip=arsip)


@app.route('/arsip/file/<int:aid>')
@login_required
def arsip_file(aid):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM arsip WHERE id=%s AND aktif=1", (aid,))
    arsip = cur.fetchone()
    cur.close(); conn.close()
    if not arsip or not arsip['filename']:
        flash('File tidak ditemukan.', 'error')
        return redirect(url_for('arsip_user'))
    path = os.path.join(app.config['SURAT_FOLDER'], arsip['filename'])
    if not os.path.exists(path):
        flash('File tidak tersedia di server.', 'error')
        return redirect(url_for('arsip_user'))
    return send_file(path, as_attachment=False)


# ── ADMIN ARSIP ───────────────────────────────────────────────────────────────

@app.route('/admin/arsip', methods=['GET'])
@admin_required
def admin_arsip():
    conn = get_db(); cur = q(conn)
    jenis_filter    = request.args.get('jenis', '')
    kategori_filter = request.args.get('kategori', '')
    q_str           = request.args.get('q', '')
    query = """SELECT a.*, u.nama as pembuat_nama
        FROM arsip a JOIN users u ON a.dibuat_oleh=u.id
        WHERE a.aktif=1"""
    params = []
    if jenis_filter:
        query += " AND a.jenis=%s"; params.append(jenis_filter)
    if kategori_filter:
        query += " AND a.kategori=%s"; params.append(kategori_filter)
    if q_str:
        query += " AND (a.judul ILIKE %s OR a.nomor ILIKE %s OR a.keterangan ILIKE %s)"
        params += [f'%{q_str}%'] * 3
    query += " ORDER BY a.tanggal DESC, a.created_at DESC"
    cur.execute(query, params)
    arsip_list = cur.fetchall()
    cur.execute("SELECT jenis, COUNT(*) as total FROM arsip WHERE aktif=1 GROUP BY jenis ORDER BY total DESC")
    stats = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/arsip.html', arsip_list=arsip_list,
        jenis_list=JENIS_ARSIP, kategori_list=KATEGORI_ARSIP,
        jenis_filter=jenis_filter, kategori_filter=kategori_filter,
        q=q_str, stats=stats, today=date.today())


@app.route('/admin/arsip/tambah', methods=['POST'])
@admin_required
def admin_arsip_tambah():
    uid = session['user_id']
    judul      = request.form.get('judul', '').strip()
    jenis      = request.form.get('jenis', '').strip()
    nomor      = request.form.get('nomor', '').strip()
    tanggal    = request.form.get('tanggal', date.today().isoformat())
    keterangan = request.form.get('keterangan', '').strip()
    if not judul or not jenis:
        flash('Judul dan jenis wajib diisi.', 'error')
        return redirect(url_for('admin_arsip'))
    filename = None; original_name = None
    f = request.files.get('file')
    if f and f.filename:
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext not in ALLOWED_DOSIR:
            flash('Format file tidak diizinkan. Gunakan PDF, JPG, atau PNG.', 'error')
            return redirect(url_for('admin_arsip'))
        os.makedirs(app.config['SURAT_FOLDER'], exist_ok=True)
        original_name = f.filename
        filename = secure_filename(f"arsip_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
        f.save(os.path.join(app.config['SURAT_FOLDER'], filename))
    conn = get_db(); cur = q(conn)
    cur.execute("""INSERT INTO arsip
        (judul, jenis, nomor, tanggal, keterangan, filename, original_name, dibuat_oleh, sumber)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'admin') RETURNING id""",
        (judul, jenis, nomor or None, tanggal, keterangan or None, filename, original_name, uid))
    aid = cur.fetchone()['id']
    log_audit(conn, 'CREATE', 'arsip', deskripsi=f'Tambah arsip: {judul}', ref_id=aid, ref_table='arsip')
    conn.commit(); cur.close(); conn.close()
    flash(f'Arsip "{judul}" berhasil ditambahkan.', 'success')
    # Jika jenis bisa dibagikan, redirect ke halaman bagikan
    if jenis in JENIS_BISA_BAGIKAN:
        return redirect(url_for('admin_arsip_bagikan', aid=aid))
    return redirect(url_for('admin_arsip'))


@app.route('/admin/arsip/bagikan/<int:aid>', methods=['GET', 'POST'])
@admin_required
def admin_arsip_bagikan(aid):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM arsip WHERE id=%s AND aktif=1", (aid,))
    arsip = cur.fetchone()
    if not arsip:
        flash('Arsip tidak ditemukan.', 'error')
        cur.close(); conn.close()
        return redirect(url_for('admin_arsip'))

    if request.method == 'POST':
        user_ids = request.form.getlist('user_ids')
        if not user_ids:
            flash('Pilih minimal satu penerima.', 'error')
            cur.close(); conn.close()
            return redirect(url_for('admin_arsip_bagikan', aid=aid))
        cfg = _get_notif_config(conn)
        berhasil = gagal = 0
        for uid_str in user_ids:
            try:
                uid_recv = int(uid_str)
                # Simpan ke arsip_bagikan (skip jika sudah ada)
                cur.execute("""INSERT INTO arsip_bagikan (arsip_id, user_id)
                    VALUES (%s, %s) ON CONFLICT (arsip_id, user_id) DO NOTHING""", (aid, uid_recv))
                # Ambil data user
                cur.execute("SELECT nama, no_hp FROM users WHERE id=%s", (uid_recv,))
                usr = cur.fetchone()
                if not usr:
                    continue
                # Kirim notifikasi in-app
                kirim_notif(conn, uid_recv,
                    judul=f"📋 {arsip['jenis']}: {arsip['judul']}",
                    pesan=f"Anda mendapat distribusi {arsip['jenis']}. Silakan cek menu Arsip.",
                    tipe='info', ref_id=aid, ref_type='arsip')
                # Kirim WA jika punya nomor
                wa_ok = False; wa_err = ''
                if usr.get('no_hp'):
                    wa_ok, wa_err = _kirim_wa_arsip(cfg, usr['no_hp'], usr['nama'], arsip['judul'], arsip['jenis'])
                cur.execute("""UPDATE arsip_bagikan SET wa_sent=%s, wa_error=%s
                    WHERE arsip_id=%s AND user_id=%s""",
                    (1 if wa_ok else 0, wa_err or None, aid, uid_recv))
                if wa_ok: berhasil += 1
                else: gagal += 1
            except Exception:
                gagal += 1
        conn.commit()
        log_audit(conn, 'UPLOAD', 'arsip', deskripsi=f'Bagikan arsip id={aid} ke {len(user_ids)} user', ref_id=aid, ref_table='arsip')
        cur.close(); conn.close()
        flash(f'Surat dibagikan ke {len(user_ids)} penerima. WA terkirim: {berhasil}, gagal: {gagal}.', 'success')
        return redirect(url_for('admin_arsip'))

    # GET — tampilkan form pilih penerima
    cur.execute("SELECT id, nama, nik, jabatan, departemen, no_hp FROM users WHERE status='active' ORDER BY departemen, nama")
    semua_user = cur.fetchall()
    # Siapa yang sudah pernah dibagikan
    cur.execute("SELECT user_id FROM arsip_bagikan WHERE arsip_id=%s", (aid,))
    sudah_bagikan = {r['user_id'] for r in cur.fetchall()}
    cur.close(); conn.close()
    return render_template('admin/arsip_bagikan.html',
        arsip=arsip, semua_user=semua_user, sudah_bagikan=sudah_bagikan)


@app.route('/admin/arsip/edit/<int:aid>', methods=['POST'])
@admin_required
def admin_arsip_edit(aid):
    uid = session['user_id']
    judul      = request.form.get('judul','').strip()
    jenis      = request.form.get('jenis','').strip()
    nomor      = request.form.get('nomor','').strip()
    tanggal    = request.form.get('tanggal', date.today().isoformat())
    keterangan = request.form.get('keterangan','').strip()
    conn = get_db(); cur = q(conn)
    f = request.files.get('file')
    if f and f.filename:
        ext = f.filename.rsplit('.',1)[-1].lower() if '.' in f.filename else ''
        if ext not in ALLOWED_DOSIR:
            flash('Format file tidak diizinkan.', 'error')
            cur.close(); conn.close()
            return redirect(url_for('admin_arsip'))
        os.makedirs(app.config['SURAT_FOLDER'], exist_ok=True)
        original_name = f.filename
        filename = secure_filename(f"arsip_{uid}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{ext}")
        f.save(os.path.join(app.config['SURAT_FOLDER'], filename))
        cur.execute("""UPDATE arsip SET judul=%s,jenis=%s,nomor=%s,tanggal=%s,keterangan=%s,
            filename=%s,original_name=%s WHERE id=%s""",
            (judul, jenis, nomor or None, tanggal, keterangan or None, filename, original_name, aid))
    else:
        cur.execute("""UPDATE arsip SET judul=%s,jenis=%s,nomor=%s,tanggal=%s,keterangan=%s
            WHERE id=%s""",
            (judul, jenis, nomor or None, tanggal, keterangan or None, aid))
    log_audit(conn, 'UPDATE', 'arsip', deskripsi=f'Edit arsip id={aid}', ref_id=aid, ref_table='arsip')
    conn.commit(); cur.close(); conn.close()
    flash('Arsip berhasil diperbarui.', 'success')
    return redirect(url_for('admin_arsip'))


@app.route('/admin/arsip/hapus/<int:aid>', methods=['POST'])
@admin_required
def admin_arsip_hapus(aid):
    conn = get_db(); cur = q(conn)
    cur.execute("UPDATE arsip SET aktif=0 WHERE id=%s", (aid,))
    log_audit(conn, 'DELETE', 'arsip', deskripsi=f'Hapus arsip id={aid}', ref_id=aid, ref_table='arsip')
    conn.commit(); cur.close(); conn.close()
    flash('Arsip berhasil dihapus.', 'success')
    return redirect(url_for('admin_arsip'))


@app.route('/admin/arsip/file/<int:aid>')
@admin_required
def admin_arsip_file(aid):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM arsip WHERE id=%s", (aid,))
    arsip = cur.fetchone()
    cur.close(); conn.close()
    if not arsip or not arsip['filename']:
        flash('File tidak ditemukan.', 'error')
        return redirect(url_for('admin_arsip'))
    path = os.path.join(app.config['SURAT_FOLDER'], arsip['filename'])
    return send_file(path, as_attachment=False)

# ── MASTER ROLE ───────────────────────────────────────────────────────────────
@app.route('/admin/master-role', methods=['GET','POST'])
@admin_required
def admin_master_role():
    conn = get_db(); cur = q(conn)
    if request.method == 'POST':
        aksi = request.form.get('aksi')
        if aksi == 'tambah':
            kode  = request.form.get('kode','').strip().lower().replace(' ','_')
            nama  = request.form.get('nama','').strip()
            desk  = request.form.get('deskripsi','').strip()
            urut  = int(request.form.get('urutan', 99))
            try:
                cur.execute("INSERT INTO master_role (kode,nama,deskripsi,urutan) VALUES (%s,%s,%s,%s)",
                    (kode, nama, desk, urut))
                conn.commit()
                flash(f'Role "{nama}" berhasil ditambahkan.', 'success')
            except Exception:
                conn.rollback()
                flash('Kode role sudah ada.', 'error')
        elif aksi == 'edit':
            rid  = request.form.get('id')
            nama = request.form.get('nama','').strip()
            desk = request.form.get('deskripsi','').strip()
            urut = int(request.form.get('urutan', 99))
            aktif= int(request.form.get('aktif', 1))
            cur.execute("UPDATE master_role SET nama=%s,deskripsi=%s,urutan=%s,aktif=%s WHERE id=%s",
                (nama, desk, urut, aktif, rid))
            conn.commit()
            flash('Role berhasil diperbarui.', 'success')
        elif aksi == 'hapus':
            rid = request.form.get('id')
            cur.execute("SELECT COUNT(*) as c FROM users WHERE role=(SELECT kode FROM master_role WHERE id=%s)", (rid,))
            cnt = cur.fetchone()['c']
            if cnt > 0:
                flash(f'Role tidak bisa dihapus, masih dipakai {cnt} pegawai.', 'error')
            else:
                cur.execute("DELETE FROM master_role WHERE id=%s", (rid,))
                conn.commit()
                flash('Role berhasil dihapus.', 'success')
        cur.close(); conn.close()
        return redirect(url_for('admin_master_role'))
    cur.execute("SELECT mr.*, COUNT(u.id) as jml_user FROM master_role mr LEFT JOIN users u ON u.role=mr.kode GROUP BY mr.id ORDER BY mr.urutan")
    roles = cur.fetchall()
    cur.close(); conn.close()
    return render_template('admin/master_role.html', roles=roles)

# ── REORDER ROLE ──────────────────────────────────────────────────────────────
@app.route('/admin/master-role/reorder', methods=['POST'])
@admin_required
def admin_master_role_reorder():
    try:
        data = request.get_json()
        order = data.get('order', [])
        conn = get_db(); cur = q(conn)
        for idx, rid in enumerate(order):
            cur.execute("UPDATE master_role SET urutan=%s WHERE id=%s", (idx, rid))
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

# ── HAK AKSES ROLE ────────────────────────────────────────────────────────────

def get_role_permissions(role_kode):
    """Kembalikan set modul_kode yang diizinkan untuk role tertentu."""
    try:
        conn = get_db(); cur = q(conn)
        cur.execute("SELECT modul_kode FROM role_permission WHERE role_kode=%s AND aktif=1", (role_kode,))
        perms = {r['modul_kode'] for r in cur.fetchall()}
        cur.close(); conn.close()
        return perms
    except Exception:
        return set()

def has_permission(modul_kode):
    """Cek apakah user session punya akses ke modul_kode."""
    role = session.get('role', '')
    if role == 'admin':
        return True
    return modul_kode in get_role_permissions(role)

def permission_required(modul_kode):
    """Decorator: tolak akses jika tidak punya permission."""
    def decorator(f):
        @wraps(f)
        def dec(*a, **kw):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if not has_permission(modul_kode):
                flash('Anda tidak memiliki akses ke fitur ini.', 'error')
                return redirect(url_for('dashboard'))
            return f(*a, **kw)
        return dec
    return decorator

@app.route('/admin/role-permission')
@admin_required
def admin_role_permission():
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM master_role WHERE aktif=1 ORDER BY urutan")
    roles = cur.fetchall()
    # Ambil semua modul unik, dikelompokkan per grup (urut berdasarkan modul_kode)
    cur.execute("""SELECT DISTINCT modul_kode, modul_nama, grup
        FROM role_permission ORDER BY grup, modul_nama""")
    modul_rows = cur.fetchall()
    # Buat dict: {role_kode: {modul_kode: aktif}}
    cur.execute("SELECT role_kode, modul_kode, aktif FROM role_permission")
    perm_map = {}
    for r in cur.fetchall():
        perm_map.setdefault(r['role_kode'], {})[r['modul_kode']] = r['aktif']
    # Kelompokkan modul per grup
    from collections import OrderedDict
    grups = OrderedDict()
    for m in modul_rows:
        grups.setdefault(m['grup'], []).append(m)
    cur.close(); conn.close()
    return render_template('admin/role_permission.html',
        roles=roles, grups=grups, perm_map=perm_map)

@app.route('/admin/role-permission/save', methods=['POST'])
@admin_required
def admin_role_permission_save():
    """Simpan seluruh matrix permission dari form POST."""
    conn = get_db(); cur = q(conn)
    # Ambil semua kombinasi role × modul yang ada
    cur.execute("SELECT role_kode, modul_kode FROM role_permission")
    all_pairs = {(r['role_kode'], r['modul_kode']) for r in cur.fetchall()}
    # Semua checkbox yang dicentang dikirim sebagai perm_<role>_<modul>
    checked = set()
    for key in request.form:
        if key.startswith('perm_'):
            parts = key[5:].split('_', 1)  # perm_{role}_{modul}
            if len(parts) == 2:
                checked.add((parts[0], parts[1]))
    # Update semua
    for role_kode, modul_kode in all_pairs:
        aktif = 1 if (role_kode, modul_kode) in checked else 0
        cur.execute("UPDATE role_permission SET aktif=%s WHERE role_kode=%s AND modul_kode=%s",
            (aktif, role_kode, modul_kode))
    conn.commit(); cur.close(); conn.close()
    flash('Hak akses berhasil disimpan.', 'success')
    return redirect(url_for('admin_role_permission'))

@app.route('/admin/role-permission/api', methods=['POST'])
@admin_required
def admin_role_permission_api():
    """Toggle satu permission via AJAX (untuk toggle langsung di matrix)."""
    try:
        data = request.get_json()
        role_kode  = data['role_kode']
        modul_kode = data['modul_kode']
        aktif      = int(data['aktif'])
        conn = get_db(); cur = q(conn)
        cur.execute("""INSERT INTO role_permission (role_kode,modul_kode,modul_nama,grup,aktif)
            VALUES (%s,%s,(SELECT modul_nama FROM role_permission WHERE modul_kode=%s LIMIT 1),
                   (SELECT grup FROM role_permission WHERE modul_kode=%s LIMIT 1),%s)
            ON CONFLICT (role_kode,modul_kode) DO UPDATE SET aktif=%s""",
            (role_kode, modul_kode, modul_kode, modul_kode, aktif, aktif))
        conn.commit(); cur.close(); conn.close()
        return jsonify({'ok': True, 'aktif': aktif})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 400

@app.route('/admin/role-permission/preset/<role_kode>', methods=['POST'])
@admin_required
def admin_role_permission_preset(role_kode):
    """Terapkan preset (aktifkan semua / nonaktifkan semua) untuk satu role."""
    mode = request.json.get('mode', 'none')  # 'all' atau 'none'
    aktif = 1 if mode == 'all' else 0
    conn = get_db(); cur = q(conn)
    cur.execute("UPDATE role_permission SET aktif=%s WHERE role_kode=%s", (aktif, role_kode))
    conn.commit(); cur.close(); conn.close()
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════════════════════
# ██  ROUTE — AUDIT LOG  ███████████████████████████████████████████████████████
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/admin/audit-log')
@admin_required
def audit_log_index():
    conn = get_db(); cur = q(conn)
    page      = request.args.get('page', 1, type=int)
    per_page  = min(request.args.get('per_page', 50, type=int), 200)
    offset    = (page - 1) * per_page
    aksi      = request.args.get('aksi', '')
    modul     = request.args.get('modul', '')
    user_id_f = request.args.get('user_id', '')
    status_f  = request.args.get('status', '')
    tgl_dari  = request.args.get('tgl_dari', '')
    tgl_sampai= request.args.get('tgl_sampai', '')
    cari      = request.args.get('cari', '')

    conditions, params = [], []
    if aksi:       conditions.append("aksi=%s");            params.append(aksi)
    if modul:      conditions.append("modul=%s");           params.append(modul)
    if user_id_f and user_id_f.isdigit():
                   conditions.append("user_id=%s");         params.append(int(user_id_f))
    if status_f:   conditions.append("status=%s");          params.append(status_f)
    if tgl_dari:   conditions.append("waktu::date>=%s");    params.append(tgl_dari)
    if tgl_sampai: conditions.append("waktu::date<=%s");    params.append(tgl_sampai)
    if cari:
        conditions.append("(deskripsi ILIKE %s OR user_nama ILIKE %s OR pesan_error ILIKE %s)")
        like = f'%{cari}%'; params += [like,like,like]

    where = ('WHERE ' + ' AND '.join(conditions)) if conditions else ''
    cur.execute(f"SELECT COUNT(*) as c FROM audit_log {where}", params)
    total = cur.fetchone()['c']
    total_pages = max(1, (total + per_page - 1) // per_page)

    cur.execute(f"""SELECT al.* FROM audit_log al
        {where} ORDER BY al.waktu DESC LIMIT %s OFFSET %s""",
        params + [per_page, offset])
    logs = cur.fetchall()

    cur.execute("""SELECT
        COUNT(*) FILTER (WHERE status='success') as sukses,
        COUNT(*) FILTER (WHERE status='error')   as error,
        COUNT(DISTINCT user_id)                  as user_aktif,
        COUNT(*) FILTER (WHERE aksi='LOGIN')     as total_login,
        COUNT(*) FILTER (WHERE aksi='LOGIN_GAGAL') as login_gagal,
        COUNT(*) FILTER (WHERE aksi IN ('CREATE','UPDATE','DELETE')) as mutasi
        FROM audit_log WHERE waktu >= NOW() - INTERVAL '24 hours'""")
    stats = cur.fetchone()

    cur.execute("""SELECT date_trunc('hour',waktu) as jam, COUNT(*) as jumlah
        FROM audit_log WHERE waktu >= NOW() - INTERVAL '12 hours'
        GROUP BY jam ORDER BY jam""")
    aktivitas_per_jam = [dict(r) for r in cur.fetchall()]

    cur.execute("""SELECT user_nama, user_role, COUNT(*) as aksi_count
        FROM audit_log WHERE waktu::date=CURRENT_DATE AND user_id IS NOT NULL
        GROUP BY user_nama, user_role ORDER BY aksi_count DESC LIMIT 10""")
    top_users = cur.fetchall()

    cur.execute("""SELECT DISTINCT user_id, user_nama, user_role
        FROM audit_log WHERE user_id IS NOT NULL ORDER BY user_nama LIMIT 100""")
    daftar_user = cur.fetchall()

    cur.close(); conn.close()
    return render_template('admin/audit_log.html',
        logs=logs, stats=stats, aktivitas_per_jam=aktivitas_per_jam,
        top_users=top_users, daftar_user=daftar_user,
        total=total, total_pages=total_pages, page=page, per_page=per_page,
        aksi=aksi, modul=modul, user_id=user_id_f,
        status=status_f, tgl_dari=tgl_dari, tgl_sampai=tgl_sampai, cari=cari,
        AKSI_LABELS=AKSI_LABELS, MODUL_LABELS=MODUL_LABELS)


@app.route('/admin/audit-log/detail/<int:log_id>')
@admin_required
def audit_log_detail(log_id):
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT * FROM audit_log WHERE id=%s", (log_id,))
    log = cur.fetchone()
    cur.close(); conn.close()
    if not log:
        flash('Log tidak ditemukan.', 'error')
        return redirect(url_for('audit_log_index'))
    return render_template('admin/audit_log_detail.html',
        log=dict(log), AKSI_LABELS=AKSI_LABELS, MODUL_LABELS=MODUL_LABELS)


@app.route('/admin/audit-log/api-stats')
@admin_required
def audit_log_api_stats():
    jam = request.args.get('jam', 24, type=int)
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT date_trunc('hour',waktu) as jam,
        COUNT(*) as total,
        COUNT(*) FILTER (WHERE status='error') as error
        FROM audit_log WHERE waktu >= NOW() - INTERVAL '%s hours'
        GROUP BY jam ORDER BY jam""", (jam,))
    per_jam = [dict(r) for r in cur.fetchall()]
    cur.execute("""SELECT aksi, COUNT(*) as jumlah FROM audit_log
        WHERE waktu >= NOW() - INTERVAL '24 hours'
        GROUP BY aksi ORDER BY jumlah DESC LIMIT 10""")
    per_aksi = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(per_jam=per_jam, per_aksi=per_aksi)


@app.route('/admin/audit-log/export')
@admin_required
def audit_log_export():
    import csv
    tgl_dari   = request.args.get('tgl_dari',   '').strip() or (date.today()-timedelta(days=30)).isoformat()
    tgl_sampai = request.args.get('tgl_sampai', '').strip() or date.today().isoformat()
    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT waktu,user_nama,user_role,aksi,modul,deskripsi,
        ref_id,ref_table,ip_address,status,pesan_error
        FROM audit_log WHERE waktu::date BETWEEN %s AND %s ORDER BY waktu DESC""",
        (tgl_dari, tgl_sampai))
    rows = cur.fetchall(); cur.close(); conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Waktu','User','Role','Aksi','Modul','Deskripsi',
                     'Ref ID','Ref Table','IP Address','Status','Error'])
    for r in rows:
        writer.writerow([r['waktu'],r['user_nama'],r['user_role'],
            r['aksi'],r['modul'],r['deskripsi'],r['ref_id'],r['ref_table'],
            r['ip_address'],r['status'],r['pesan_error']])
    from flask import Response
    return Response('\ufeff'+output.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition':f'attachment; filename=audit_log_{tgl_dari}_{tgl_sampai}.csv'})


@app.route('/admin/audit-log/purge', methods=['POST'])
@admin_required
def audit_log_purge():
    hari = max(request.form.get('hari', 90, type=int), 7)
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM audit_log WHERE waktu < NOW() - INTERVAL '%s days'", (hari,))
    deleted = cur.rowcount; conn.commit(); cur.close()
    log_audit(conn, 'DELETE', 'sistem',
        deskripsi=f'Purge audit log > {hari} hari — {deleted} baris dihapus')
    conn.close()
    flash(f'Berhasil menghapus {deleted} baris log lama (>{hari} hari).', 'success')
    return redirect(url_for('audit_log_index'))


# ══════════════════════════════════════════════════════════════════════════════
# ██  ROUTE — LUPA PASSWORD  ███████████████████████████████████████████████████
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/lupa-password', methods=['GET','POST'])
def lupa_password():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'GET':
        return render_template('lupa_password.html', step='input')

    identitas = request.form.get('identitas', '').strip()
    metode    = request.form.get('metode', 'email')
    if not identitas:
        flash('Masukkan email atau nomor WhatsApp Anda.', 'error')
        return render_template('lupa_password.html', step='input')

    conn = get_db(); cur = q(conn)
    cur.execute("""SELECT id,nama,email,no_hp,status FROM users
        WHERE email=%s OR no_hp=%s LIMIT 1""", (identitas, identitas))
    user = cur.fetchone()

    if not user:
        cur.close(); conn.close()
        log_audit(conn, 'OTP_KIRIM', 'auth',
            deskripsi=f'Lupa password — akun tidak ditemukan: {identitas}',
            status='error', user_id=None, user_nama=identitas, user_role='-')
        flash('Jika akun dengan data tersebut ada, OTP akan dikirimkan segera.', 'info')
        return render_template('lupa_password.html', step='input')

    if user['status'] in ('pending','rejected'):
        cur.close(); conn.close()
        flash('Akun belum aktif. Hubungi administrator.', 'warning')
        return render_template('lupa_password.html', step='input')

    user = dict(user)
    cfg  = _get_notif_config(conn)
    if metode == 'whatsapp':
        tujuan = user.get('no_hp') or ''
        if not tujuan:
            cur.close(); conn.close()
            flash('Nomor WhatsApp tidak terdaftar pada akun ini. Coba via Email.', 'warning')
            return render_template('lupa_password.html', step='input')
    else:
        metode = 'email'
        tujuan = user.get('email') or ''

    otp = _generate_otp(cfg['otp_length'])
    _simpan_otp(conn, user['id'], otp, metode, tujuan, cfg['otp_expire'])
    cur.close()

    ok, err_msg = (_kirim_email_otp if metode=='email' else _kirim_wa_otp)(cfg, tujuan, user['nama'], otp)
    if not ok:
        conn.close()
        log_audit(conn, 'OTP_KIRIM', 'auth',
            deskripsi=f'Gagal kirim OTP {metode}: {err_msg}',
            status='error', ref_id=user['id'], ref_table='users')
        flash(f'Gagal mengirim OTP: {err_msg}', 'error')
        return render_template('lupa_password.html', step='input')

    log_audit(conn, 'OTP_KIRIM', 'auth',
        deskripsi=f'Kirim OTP {metode} ke {_mask_tujuan(tujuan, metode)}',
        ref_id=user['id'], ref_table='users',
        user_id=user['id'], user_nama=user['nama'], user_role='-')
    conn.close()

    session['reset_user_id']  = user['id']
    session['reset_metode']   = metode
    session['reset_tujuan']   = _mask_tujuan(tujuan, metode)
    session['reset_ts']       = datetime.now().isoformat()
    flash(f'OTP telah dikirim ke {_mask_tujuan(tujuan, metode)}.', 'success')
    return render_template('lupa_password.html', step='otp',
        metode=metode, tujuan=_mask_tujuan(tujuan, metode), expire=cfg['otp_expire'])


@app.route('/lupa-password/verifikasi', methods=['POST'])
def verifikasi_otp():
    user_id = session.get('reset_user_id')
    if not user_id:
        flash('Sesi reset telah berakhir. Mulai ulang.', 'error')
        return redirect(url_for('lupa_password'))
    otp_input = request.form.get('otp', '').strip()
    if len(otp_input) < 4:
        flash('Kode OTP tidak valid.', 'error')
        return render_template('lupa_password.html', step='otp',
            metode=session.get('reset_metode'), tujuan=session.get('reset_tujuan'))
    conn = get_db()
    row  = _verifikasi_otp(conn, user_id, otp_input)
    if not row:
        log_audit(conn, 'OTP_VERIF', 'auth',
            deskripsi=f'OTP salah/kadaluarsa untuk user_id {user_id}',
            status='error', ref_id=user_id, ref_table='users')
        conn.close()
        flash('Kode OTP salah atau sudah kadaluarsa. Coba kirim ulang.', 'error')
        return render_template('lupa_password.html', step='otp',
            metode=session.get('reset_metode'), tujuan=session.get('reset_tujuan'))
    log_audit(conn, 'OTP_VERIF', 'auth',
        deskripsi=f'OTP berhasil diverifikasi untuk user_id {user_id}',
        ref_id=user_id, ref_table='users')
    conn.close()
    session['reset_otp_id']   = row['id']
    session['reset_verified'] = True
    return render_template('lupa_password.html', step='reset')


@app.route('/lupa-password/reset', methods=['POST'])
def reset_password():
    user_id  = session.get('reset_user_id')
    otp_id   = session.get('reset_otp_id')
    verified = session.get('reset_verified')
    if not (user_id and otp_id and verified):
        flash('Sesi reset tidak valid. Mulai ulang.', 'error')
        return redirect(url_for('lupa_password'))
    pw_baru    = request.form.get('password_baru', '')
    pw_konfirm = request.form.get('password_konfirm', '')
    if len(pw_baru) < 6:
        flash('Password minimal 6 karakter.', 'error')
        return render_template('lupa_password.html', step='reset')
    if pw_baru != pw_konfirm:
        flash('Konfirmasi password tidak cocok.', 'error')
        return render_template('lupa_password.html', step='reset')
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE users SET password=%s WHERE id=%s",
        (generate_password_hash(pw_baru), user_id))
    conn.commit()
    _tandai_otp_digunakan(conn, otp_id)
    log_audit(conn, 'RESET_PW', 'auth',
        deskripsi=f'Password berhasil direset untuk user_id {user_id}',
        ref_id=user_id, ref_table='users')
    cur.close(); conn.close()
    for k in ['reset_user_id','reset_otp_id','reset_verified',
              'reset_metode','reset_tujuan','reset_ts']:
        session.pop(k, None)
    flash('Password berhasil direset! Silakan login dengan password baru Anda.', 'success')
    return redirect(url_for('login'))


@app.route('/lupa-password/kirim-ulang', methods=['POST'])
def kirim_ulang_otp():
    user_id = session.get('reset_user_id')
    metode  = session.get('reset_metode', 'email')
    if not user_id:
        return jsonify(success=False, message='Sesi tidak valid.')
    conn = get_db(); cur = q(conn)
    cur.execute("SELECT nama,email,no_hp FROM users WHERE id=%s", (user_id,))
    user = cur.fetchone()
    if not user: cur.close(); conn.close(); return jsonify(success=False, message='User tidak ditemukan.')
    user = dict(user)
    cfg  = _get_notif_config(conn)
    tujuan = user.get('no_hp') if metode=='whatsapp' else user.get('email')
    otp = _generate_otp(cfg['otp_length'])
    _simpan_otp(conn, user_id, otp, metode, tujuan, cfg['otp_expire'])
    ok, err = (_kirim_email_otp if metode=='email' else _kirim_wa_otp)(cfg, tujuan, user['nama'], otp)
    log_audit(conn, 'OTP_KIRIM', 'auth',
        deskripsi=f'Kirim ulang OTP {metode} — {"sukses" if ok else "gagal: "+err}',
        ref_id=user_id, ref_table='users', status='success' if ok else 'error')
    cur.close(); conn.close()
    if ok: return jsonify(success=True, message=f'OTP baru dikirim ke {_mask_tujuan(tujuan,metode)}.')
    return jsonify(success=False, message=f'Gagal kirim ulang: {err}')


@app.route('/admin/settings/notif', methods=['POST'])
@admin_required
def admin_settings_notif():
    conn = get_db(); cur = conn.cursor()
    try:
        cur.execute("""UPDATE settings SET
            smtp_host=%s, smtp_port=%s, smtp_user=%s, smtp_pass=%s,
            smtp_from_name=%s, smtp_tls=%s, fonnte_token=%s WHERE id=1""",
            (request.form.get('smtp_host','').strip(),
             int(request.form.get('smtp_port',587) or 587),
             request.form.get('smtp_user','').strip(),
             request.form.get('smtp_pass','').strip(),
             request.form.get('smtp_from_name','Presensi Digital').strip(),
             request.form.get('smtp_tls','true')=='true',
             request.form.get('fonnte_token','').strip()))
        conn.commit()
        log_audit(conn, 'SETTING', 'settings', deskripsi='Update konfigurasi SMTP & WhatsApp')
        flash('Konfigurasi notifikasi berhasil disimpan!', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Gagal menyimpan: {e}', 'error')
    finally:
        cur.close(); conn.close()
    return redirect(url_for('admin_settings'))


@app.route('/admin/test-notif', methods=['POST'])
@admin_required
def admin_test_notif():
    metode = request.form.get('metode', 'email')
    tujuan = request.form.get('tujuan', '').strip()
    if not tujuan: return jsonify(success=False, message='Tujuan tidak boleh kosong')
    conn = get_db(); cfg = _get_notif_config(conn); conn.close()
    otp = _generate_otp(cfg['otp_length'])
    nama = session.get('nama', 'Admin')
    ok, err = (_kirim_email_otp if metode=='email' else _kirim_wa_otp)(cfg, tujuan, nama, otp)
    if ok: return jsonify(success=True, message=f'Test OTP ({otp}) berhasil dikirim ke {tujuan}')
    return jsonify(success=False, message=err)


# Jalankan init_db saat modul dimuat (termasuk saat dijalankan via Gunicorn)
with app.app_context():
    try:
        init_db()
    except Exception as _e:
        print(f"[init_db] warning: {_e}")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5030)